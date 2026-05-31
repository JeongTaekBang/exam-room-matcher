"""강의실 배정 프로그램 대시보드

시간표와 요청사항을 시각화하여 사람의 수동 배정을 돕는다.
충돌 감지, 잔여 슬롯 히트맵, 클릭 기반 배정 기능 포함.
"""

import html as html_mod
import io
import re
import time
import argparse
import os
from pathlib import Path
from collections import Counter, defaultdict

import openpyxl as _openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st
import pandas as pd

from data_loader import (
    load_all, Category,
    DATE_TO_DAY as _DEFAULT_DATE_TO_DAY,
    DAY_TO_SHEET as _DEFAULT_DAY_TO_SHEET,
    SHEET_ORDER as _DEFAULT_SHEET_ORDER,
)
from workflow_utils import (
    StaleFileError,
    append_audit_event,
    collect_extra_room_slots,
    load_assignments,
    load_releases,
    read_audit_events,
    releases_to_slot_set,
    resolve_exam_room,
    resolve_needed_periods,
    save_assignments,
    save_releases,
)
from conflict_check import (
    detect_manual_processed_entries,
    detect_processed_room_conflicts,
    detect_timetable_overlaps,
)

ROOT = Path(__file__).resolve().parent
ASSIGNMENTS_FILENAME = "_assignments.json"
RELEASES_FILENAME = "_releases.json"
AUDIT_LOG_FILENAME = "_assignment_audit.jsonl"

st.set_page_config(page_title="강의실 배정 프로그램", layout="wide")

# ── Streamlit 버전 호환 ──
# 1.49+는 st.dataframe(**_STRETCH)를 받지만, 그 이전 버전(예: run.bat이
# 실행하는 conda base의 1.47)은 width에 정수만 허용해 "stretch" 문자열에서
# TypeError가 난다. 설치된 버전에 맞춰 인자를 자동 선택한다.
def _st_version_tuple():
    parts = []
    for chunk in st.__version__.split(".")[:3]:
        digits = ""
        for ch in chunk:
            if ch.isdigit():
                digits += ch
            else:
                break
        parts.append(int(digits) if digits else 0)
    return tuple(parts)


_STRETCH = (
    {"width": "stretch"}
    if _st_version_tuple() >= (1, 49)
    else {"use_container_width": True}
)

from assignment_status import (
    CAT_LABELS,
    LABEL_NORMAL_EXAM,
    LABEL_NO_EXAM,
    LABEL_ROOM_CHANGE,
    LABEL_ROOM_SPLIT,
    LABEL_SKIP,
    _room_cap,
    compute_auto_released,
    compute_status,
    compute_unused_rooms,
    extract_base_key,
    resolve_processed_room,
)
CAT_COLORS = {
    Category.NORMAL_EXAM: "#4472C4",
    Category.NO_EXAM: "#FF69B4",
    Category.ROOM_CHANGE: "#FFD700",
    Category.ROOM_SPLIT: "#9370DB",
    Category.SKIP: "#AAAAAA",
}
DARK_BG = ("#4472C4", "#D2691E", "#5F9EA0", "#9370DB", "#FF0000", "#228B22")

# compute_status / _room_cap / CAT_LABELS·LABEL_* 는 assignment_status.py 로 분리
# (Streamlit 의존성 없이 단위 테스트 가능하도록).


# ──────────────────────────────────────────────
# 유틸
# ──────────────────────────────────────────────

def scan_folders():
    entries = []
    for year_dir in sorted(ROOT.iterdir()):
        if not year_dir.is_dir() or not year_dir.name.isdigit():
            continue
        for exam_dir in sorted(year_dir.iterdir()):
            if not exam_dir.is_dir():
                continue
            xlsx = [f for f in exam_dir.iterdir()
                    if f.suffix == ".xlsx" and not f.name.startswith("~$")]
            req = next((f for f in xlsx if "요청사항" in f.name), None)
            tt = next((f for f in xlsx if "타임테이블" in f.name), None)
            if req and tt:
                entries.append((f"{year_dir.name} / {exam_dir.name}", exam_dir, req, tt))
    return entries


@st.cache_data(show_spinner="데이터 로딩 중...")
def cached_load(req_path: str, tt_path: str, req_mtime: float = 0, tt_mtime: float = 0):
    # req_mtime/tt_mtime은 캐시 키 일부 — 엑셀 파일이 바뀌면(요청사항/P열 수정 등)
    # mtime이 달라져 캐시가 자동 무효화된다. (밑줄 접두 인자는 st.cache_data가
    # 해시에서 제외하므로 절대 밑줄을 붙이면 안 된다.)
    return load_all(req_path, tt_path)


def generate_request_with_processed(req_path: str, requests, assignments) -> bytes:
    """원본 요청 엑셀의 P열(16열, '요청사항 처리')에 프로그램 배정 결과를 채운
    사본을 bytes로 반환한다.

    배정/시험/미실시로 **확정된 행만** P열을 덮어쓰고, 미확정 행(미배정 변경·분반,
    미확정)은 원본 셀 값을 그대로 보존해 사용자가 직접 입력할 여지를 남긴다.
    다운로드용 사본만 만들며 원본 파일은 변경하지 않는다.
    """
    wb = _openpyxl.load_workbook(req_path)
    ws = wb[wb.sheetnames[0]]  # load_requests와 동일하게 첫 시트(공통) 기준
    for req in requests:
        value = resolve_processed_room(req, assignments)
        if value:  # 빈 결과는 기존 값 보존
            ws.cell(row=req.row, column=16, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def render_timetable_html(room_data: dict, color_map: dict, room_capacity: dict,
                          tooltip_map: dict = None) -> str:
    rooms = sorted(room_data.keys())
    html = ['<div style="overflow-x:auto;max-height:600px;overflow-y:auto;">'
            '<table style="border-collapse:collapse;font-size:11px;width:100%;">']
    html.append('<thead><tr style="position:sticky;top:0;z-index:1;">')
    html.append('<th style="background:#333;color:#fff;padding:3px 6px;min-width:60px;'
                'position:sticky;left:0;z-index:2;">강의실</th>')
    html.append('<th style="background:#333;color:#fff;padding:3px 4px;min-width:30px;">수용</th>')
    for p in range(15):
        html.append(f'<th style="background:#333;color:#fff;padding:3px 2px;'
                     f'min-width:70px;">{p}교시<br>{8+p}~{9+p}</th>')
    html.append('</tr></thead><tbody>')

    for room in rooms:
        cap = room_capacity.get(room, "")
        html.append('<tr>')
        html.append(f'<td style="background:#f5f5f5;color:#333;font-weight:bold;'
                     f'padding:2px 4px;border:1px solid #ddd;position:sticky;left:0;z-index:1;">{room}</td>')
        html.append(f'<td style="background:#f5f5f5;color:#333;text-align:center;'
                     f'padding:2px;border:1px solid #ddd;">{cap}</td>')

        periods = room_data.get(room, {})
        for p in range(15):
            text = periods.get(p, "")
            bg = color_map.get((room, p), "#ffffff")
            fg = "#fff" if bg in DARK_BG else "#333"
            title_attr = ""
            if tooltip_map and (room, p) in tooltip_map:
                tip = html_mod.escape(tooltip_map[(room, p)])
                title_attr = f' title="{tip}"'
            display = html_mod.escape(text[:14]) if text else ""
            html.append(
                f'<td style="background:{bg};color:{fg};padding:2px;border:1px solid #ddd;'
                f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:90px;"'
                f'{title_attr}>{display}</td>')
        html.append('</tr>')

    html.append('</tbody></table></div>')
    return "\n".join(html)


def render_availability_grid(timetable_raw: dict, room_capacity: dict,
                            periods: list, min_cap: int, assignments: dict,
                            sheet_name: str, released_slots: set,
                            highlight_room: str = "",
                            extra_slots: dict = None) -> str:
    """교시별 강의실 점유 현황을 HTML 미니 격자로 렌더링한다.

    *extra_slots*: 시간표에 없는 강의실의 점유 정보.
    ``{room: {period: subject_key}}`` 형태.
    """
    assigned_slots = set()
    if assignments:
        for a in assignments.values():
            if a["sheet"] == sheet_name:
                for ap in a["periods"]:
                    assigned_slots.add((a["room"], ap))

    # 표시 교시 범위: 필요 교시 ±2 (최소 0, 최대 14)
    p_min = max(0, min(periods) - 2)
    p_max = min(14, max(periods) + 2)
    display_periods = list(range(p_min, p_max + 1))

    if extra_slots is None:
        extra_slots = {}

    # 시간표 강의실 + 시간표 외 강의실 합산
    all_rooms = set(r for r in timetable_raw if room_capacity.get(r, 0) >= min_cap)
    all_rooms |= set(extra_slots.keys())
    rooms = sorted(all_rooms, key=lambda r: room_capacity.get(r, 0))
    if not rooms:
        return ""

    lines = ['<div style="overflow-x:auto;"><table style="border-collapse:collapse;'
             'font-size:11px;width:100%;"><thead><tr>'
             '<th style="padding:2px 4px;border:1px solid #555;background:#333;color:#fff;">강의실</th>'
             '<th style="padding:2px 4px;border:1px solid #555;background:#333;color:#fff;">수용</th>']
    for p in display_periods:
        bg = "#444" if p in periods else "#333"
        lines.append(f'<th style="padding:2px 4px;border:1px solid #555;background:{bg};'
                     f'color:#fff;">{p}</th>')
    lines.append('</tr></thead><tbody>')

    for room in rooms:
        cap = room_capacity.get(room, 0)
        is_extra = room not in timetable_raw
        row_bg = "#1a3a1a" if room == highlight_room else ""
        room_label = f"*{room}" if is_extra else room
        lines.append(f'<tr style="background:{row_bg};">'
                     f'<td style="padding:2px 4px;border:1px solid #555;font-weight:bold;'
                     f'{"color:#228B22;" if is_extra else ""}">{room_label}</td>'
                     f'<td style="padding:2px 4px;border:1px solid #555;text-align:center;">{cap or "-"}</td>')
        for p in display_periods:
            is_released = released_slots and (sheet_name, room, p) in released_slots
            occupied_tt = not is_released and p in timetable_raw.get(room, {})
            occupied_extra = not occupied_tt and p in extra_slots.get(room, {})
            assigned = (room, p) in assigned_slots
            if assigned:
                bg, fg, txt = "#4472C4", "#fff", "배정"
            elif occupied_tt:
                bg, fg, txt = "#8B4513", "#fff", "수업"
            elif occupied_extra:
                bg, fg, txt = "#228B22", "#fff", "수업"
            else:
                if p in periods:
                    bg, fg, txt = "#228B22", "#fff", "◯"
                else:
                    bg, fg, txt = "#2a2a2a", "#888", ""
            lines.append(f'<td style="padding:2px 4px;border:1px solid #555;text-align:center;'
                         f'background:{bg};color:{fg};">{txt}</td>')
        lines.append('</tr>')
    lines.append('</tbody></table></div>')
    return "".join(lines)


def get_free_rooms(timetable_raw: dict, room_capacity: dict,
                   periods: list, min_cap: int = 0,
                   assignments: dict = None, sheet_name: str = "",
                   released_slots: set = None) -> list:
    """주어진 교시 범위에서 빈 강의실 목록 반환."""
    result = []
    # 배정된 슬롯 수집
    assigned_slots = set()
    if assignments:
        for a in assignments.values():
            if a["sheet"] == sheet_name:
                for ap in a["periods"]:
                    assigned_slots.add((a["room"], ap))

    for room in sorted(timetable_raw.keys()):
        cap = room_capacity.get(room, 0)
        if cap < min_cap:
            continue
        free = True
        for p in periods:
            is_released = released_slots and (sheet_name, room, p) in released_slots
            if not is_released and p in timetable_raw.get(room, {}):
                free = False
                break
            if (room, p) in assigned_slots:
                free = False
                break
        if free:
            result.append((room, cap))
    result.sort(key=lambda x: x[1])
    return result


def build_review_queue_rows(requests: list, timetable_data: dict, room_capacity: dict,
                            assignments: dict, released_slots: set = None) -> list[dict]:
    """검수 큐(예외/수동 확인 필요 항목) 목록을 생성한다."""
    rows = []
    for req in requests:
        if req.category == Category.SKIP:
            rows.append({
                "분류": CAT_LABELS[req.category],
                "과목명": req.key,
                "이슈": "분류 미확정",
                "세부": req.skip_reason or "확인 필요",
            })
            continue

        if req.category not in (Category.NORMAL_EXAM, Category.ROOM_CHANGE, Category.ROOM_SPLIT):
            continue

        if req.exam_date not in DATE_TO_DAY:
            rows.append({
                "분류": CAT_LABELS[req.category],
                "과목명": req.key,
                "이슈": "시험일자 누락/범위 밖",
                "세부": str(req.exam_date or ""),
            })
            continue

        exam_day = DATE_TO_DAY[req.exam_date]
        sheet = _resolve_sheet(req.exam_date, exam_day)
        needed_periods = resolve_needed_periods(req, exam_day)
        if not needed_periods:
            rows.append({
                "분류": CAT_LABELS[req.category],
                "과목명": req.key,
                "이슈": "시험교시 정보 없음",
                "세부": req.schedule_raw,
            })
            continue

        day_timetable = timetable_data.get(sheet, {})

        # NORMAL_EXAM인데 시험교시 명시 없음 + 수업 2교시 이상 → 부분해제 가능성
        if req.category == Category.NORMAL_EXAM and req.exam_start is None:
            if len(needed_periods) >= 2:
                rows.append({
                    "분류": CAT_LABELS[req.category],
                    "과목명": req.key,
                    "이슈": "부분해제 확인 필요",
                    "세부": f"수업 {needed_periods}교시, 시험교시 미명시 → 수동 해제 필요",
                })

        if req.category in (Category.ROOM_CHANGE, Category.ROOM_SPLIT):
            if compute_status(req, assignments, room_capacity) == "완료":
                continue
            min_cap = 0 if req.category == Category.ROOM_SPLIT else req.students
            free = get_free_rooms(
                day_timetable,
                room_capacity,
                needed_periods,
                min_cap,
                assignments,
                sheet,
                released_slots,
            )
            if not free:
                rows.append({
                    "분류": CAT_LABELS[req.category],
                    "과목명": req.key,
                    "이슈": "배정 후보 없음",
                    "세부": f"{sheet} / {needed_periods}교시 / {req.students}명+",
                })
    return rows


def _build_day_verification(day, requests, timetable_data, assignments,
                            auto_released, released_slots):
    """주어진 일자의 결과 검증 격자를 생성한다."""
    raw = timetable_data.get(day, {})
    full_data = {}
    full_color = {}
    full_tip = {}
    conflict_map = defaultdict(list)

    for room in raw:
        full_data[room] = {}
        for p, (val, rgb) in raw[room].items():
            slot_key = (day, room, p)
            if slot_key in auto_released:
                _ar_key, _ar_label = auto_released[slot_key]
                full_data[room][p] = f"{_ar_label}: {_ar_key}"
                full_color[(room, p)] = "#FF69B4"
            elif slot_key in released_slots:
                full_data[room][p] = f"해제: {val}"
                full_color[(room, p)] = "#FF69B4"
            else:
                full_data[room][p] = val
                full_color[(room, p)] = "#5F9EA0" if rgb == "FF5F9EA0" else "#D2691E"

    # 수업시간표 슬롯 중 시간표에 없는 강의실의 점유 표시
    _weekday_m = re.search(r"\(([월화수목금토일])\)", day)
    if _weekday_m:
        _slot_weekday = _weekday_m.group(1)
        for req in requests:
            for slot in req.slots:
                if slot.day != _slot_weekday or not slot.room:
                    continue
                if slot.room in raw:
                    continue
                full_data.setdefault(slot.room, {})
                for p in range(slot.start, slot.end + 1):
                    if (slot.room, p) not in full_color:
                        full_data[slot.room][p] = f"수업: {req.key}"
                        full_color[(slot.room, p)] = "#228B22"
                        full_tip[(slot.room, p)] = f"[수업 점유] {req.key}\n시간표에 없는 강의실"

    for req in requests:
        if req.category in (Category.SKIP, Category.NO_EXAM):
            continue
        _has_any = (req.key in assignments
                    or any(k.startswith(req.key + "+") for k in assignments))
        if _has_any:
            continue
        if req.exam_date not in DATE_TO_DAY:
            continue
        exam_day = DATE_TO_DAY[req.exam_date]
        if _resolve_sheet(req.exam_date, exam_day) != day:
            continue

        cells = set()
        if req.exam_start is not None and req.exam_end is not None:
            start_p, end_p = sorted((req.exam_start, req.exam_end))
            rooms = {s.room for s in req.slots if s.day == exam_day and s.room}
            if not rooms and req.room:
                rooms = {req.room}
            for room in rooms:
                for p in range(start_p, end_p + 1):
                    cells.add((room, p))
        else:
            for slot in req.slots:
                if slot.day != exam_day or not slot.room:
                    continue
                for p in range(slot.start, slot.end + 1):
                    cells.add((slot.room, p))

        for room, p in cells:
            # 해제/자동해제된 슬롯은 오버레이하지 않음
            slot_key = (day, room, p)
            if slot_key in auto_released or slot_key in released_slots:
                continue
            full_data.setdefault(room, {})
            conflict_map[(room, p)].append(req)

    conflict_cells = set()
    for (room, p), reqs in conflict_map.items():
        # 같은 과목명의 다른 분반끼리는 시작 강의실이 같을 수 있고 둘 다
        # ROOM_CHANGE/SPLIT으로 옮겨질 예정이므로 현재 표시상 겹치는 것은
        # 실제 충돌이 아니다. 2단계 배정 충돌 감지와 동일한 형제 필터를 적용.
        subjects = {r.name for r in reqs}
        if len(reqs) > 1 and len(subjects) > 1:
            conflict_cells.add((room, p))
            full_data[room][p] = f"충돌! {len(reqs)}건"
            full_color[(room, p)] = "#FF0000"
            lines = []
            for r in reqs:
                lines.append(f"[{CAT_LABELS[r.category]}] {r.key} (수강 {r.students}명)")
                if r.remarks:
                    lines.append(f"  → {r.remarks[:60]}")
            full_tip[(room, p)] = "\n".join(lines)
        else:
            req = reqs[0]
            full_data[room][p] = req.key
            if room not in raw:
                full_color[(room, p)] = "#228B22"
            else:
                full_color[(room, p)] = CAT_COLORS[req.category]
            tip = f"[{CAT_LABELS[req.category]}] {req.key}\n수강생: {req.students}명"
            if room not in raw:
                tip += "\n⚠ 시간표에 없는 강의실"
            if req.remarks:
                tip += f"\n\n{req.remarks}"
            full_tip[(room, p)] = tip

    _split_orig_done = set()  # (orig, base_key) 단위 중복 방지
    for key, a in assignments.items():
        if a["sheet"] != day:
            continue
        is_split = a.get("category") == LABEL_ROOM_SPLIT
        base_key = extract_base_key(key)

        if is_split and a.get("keep_orig", True):
            orig = a.get("original_room", "")
            if orig and (orig, base_key) not in _split_orig_done:
                full_data.setdefault(orig, {})
                for p in a["periods"]:
                    existing = conflict_map.get((orig, p), [])
                    if not existing:
                        full_data[orig][p] = f"시험(원래): {base_key}"
                        full_color[(orig, p)] = "#4472C4"
                        full_tip[(orig, p)] = f"[원래강의실 유지] {base_key} / {orig}"
                _split_orig_done.add((orig, base_key))

        room = a["room"]
        full_data.setdefault(room, {})
        for p in a["periods"]:
            existing = conflict_map.get((room, p), [])
            if existing:
                conflict_cells.add((room, p))
                full_data[room][p] = f"충돌! 배정+{len(existing)}건"
                full_color[(room, p)] = "#FF0000"
                lines = [f"[수동배정] {key} → {room}"]
                for r in existing:
                    lines.append(f"[{CAT_LABELS[r.category]}] {r.key}")
                full_tip[(room, p)] = "\n".join(lines)
            elif is_split:
                full_data[room][p] = f"분반(추가): {key}"
                full_color[(room, p)] = "#4472C4"
                full_tip[(room, p)] = f"[분반 추가강의실] {key} / {room}"
            else:
                full_data[room][p] = f"이동: {key}"
                full_color[(room, p)] = "#4472C4"
                full_tip[(room, p)] = f"[이동 배정강의실] {key} / {room}"

    assignment_map = defaultdict(list)
    for key, a in assignments.items():
        if a["sheet"] != day:
            continue
        for p in a["periods"]:
            assignment_map[(a["room"], p)].append(key)
        # ROOM_SPLIT: 원래 강의실 유지 시에만 충돌 집계에 포함
        if a.get("category") == LABEL_ROOM_SPLIT and a.get("keep_orig", True):
            orig = a.get("original_room", "")
            if orig and orig != a["room"]:
                for p in a["periods"]:
                    assignment_map[(orig, p)].append(key)

    for (room, p), keys in assignment_map.items():
        # 같은 과목의 분반 배정끼리는 충돌이 아님 (기본키 기준 그룹핑)
        base_keys = {extract_base_key(k) for k in keys}
        if len(base_keys) <= 1:
            continue
        if len(keys) > 1:
            conflict_cells.add((room, p))
            full_data.setdefault(room, {})[p] = f"충돌! 배정{len(keys)}건"
            full_color[(room, p)] = "#FF0000"
            full_tip[(room, p)] = "\n".join(f"[수동배정] {k}" for k in keys)

    return full_data, full_color, full_tip, conflict_cells


def format_audit_details(details) -> str:
    if not isinstance(details, dict):
        return ""
    if "loaded_count" in details:
        return f"저장본 불러오기 {details['loaded_count']}건"
    if "count" in details:
        return f"전체 초기화 {details['count']}건"
    parts = []
    if details.get("sheet"):
        parts.append(str(details["sheet"]))
    if details.get("room"):
        parts.append(str(details["room"]))
    if details.get("periods"):
        parts.append(str(details["periods"]))
    return " / ".join(parts)


# ──────────────────────────────────────────────
# 사이드바
# ──────────────────────────────────────────────

entries = scan_folders()
if not entries:
    st.error("데이터 폴더가 없습니다. 년도/학기_시험 폴더에 xlsx 파일을 넣어주세요.")
    st.stop()

parser = argparse.ArgumentParser()
parser.add_argument("--folder", default=None)
args, _ = parser.parse_known_args()

default_idx = 0
if args.folder:
    for i, (_, folder, _, _) in enumerate(entries):
        if args.folder in str(folder):
            default_idx = i
            break

if "operator_name" not in st.session_state:
    st.session_state.operator_name = os.getenv("USERNAME", "")

with st.sidebar:
    st.title("강의실 배정 프로그램")
    labels = [e[0] for e in entries]
    selected = st.selectbox("데이터 선택", labels, index=default_idx)
    idx = labels.index(selected)
    _, folder, req_file, tt_file = entries[idx]
    st.caption(f"요청: {req_file.name}")
    st.caption(f"시간표: {tt_file.name}")
    st.text_input("작업자", key="operator_name", help="배정/초기화 작업 이력에 기록됩니다.")

_cur_req_mtime = req_file.stat().st_mtime
_cur_tt_mtime = tt_file.stat().st_mtime
data = cached_load(str(req_file), str(tt_file),
                   req_mtime=_cur_req_mtime,
                   tt_mtime=_cur_tt_mtime)
requests = data["requests"]
room_capacity = data["room_capacity"]
timetable_data = data["timetable_data"]
DATE_TO_DAY = data.get("date_to_day", _DEFAULT_DATE_TO_DAY)
DAY_TO_SHEET = data.get("day_to_sheet", _DEFAULT_DAY_TO_SHEET)
DAY_TO_SHEETS: dict[str, list[str]] = data.get(
    "day_to_sheets", {d: [s] for d, s in DAY_TO_SHEET.items()}
)
DATE_TO_SHEET = data.get("date_to_sheet", {})
SHEET_ORDER = data.get("sheet_order", _DEFAULT_SHEET_ORDER)
assignments_file = folder / ASSIGNMENTS_FILENAME
releases_file = folder / RELEASES_FILENAME
audit_file = folder / AUDIT_LOG_FILENAME

inv_day_sheet = {v: k for k, v in DAY_TO_SHEET.items()}

def _resolve_sheet(exam_date, exam_day=None):
    """날짜→시트 직접 매핑 우선, 폴백으로 요일→시트."""
    if exam_date and exam_date in DATE_TO_SHEET:
        return DATE_TO_SHEET[exam_date]
    if exam_day:
        return DAY_TO_SHEET.get(exam_day, "")
    if exam_date and exam_date in DATE_TO_DAY:
        return DAY_TO_SHEET.get(DATE_TO_DAY[exam_date], "")
    return ""

def _file_mtime(path: Path) -> float | None:
    return path.stat().st_mtime if path.exists() else None

def _reload_session_data():
    """배정/해제 데이터를 파일에서 다시 읽고 mtime을 갱신한다."""
    st.session_state.assignments = load_assignments(assignments_file)
    st.session_state._assignments_mtime = _file_mtime(assignments_file)
    st.session_state.releases = load_releases(releases_file)
    st.session_state._releases_mtime = _file_mtime(releases_file)
    st.session_state.released_slots = releases_to_slot_set(st.session_state.releases)

dataset_key = str(folder.resolve())
_data_changed = (st.session_state.get("_data_req_mtime") != _cur_req_mtime
                 or st.session_state.get("_data_tt_mtime") != _cur_tt_mtime)
if st.session_state.get("active_dataset_key") != dataset_key or _data_changed:
    _reload_session_data()
    st.session_state.active_dataset_key = dataset_key
    st.session_state._data_req_mtime = _cur_req_mtime
    st.session_state._data_tt_mtime = _cur_tt_mtime
elif "assignments" not in st.session_state:
    st.session_state.assignments = load_assignments(assignments_file)
    st.session_state._assignments_mtime = _file_mtime(assignments_file)
if "releases" not in st.session_state:
    st.session_state.releases = load_releases(releases_file)
    st.session_state._releases_mtime = _file_mtime(releases_file)
    st.session_state.released_slots = releases_to_slot_set(st.session_state.releases)


# ── 취소 2단계 확인 헬퍼 ──
# 운영자가 ✕ 버튼을 실수로 한 번 눌렀을 때 즉시 삭제되는 것을 막는다.
# 같은 버튼을 60초 안에 한 번 더 눌러야 실제 삭제가 일어난다.
# 동시에 하나의 pending만 허용 — 다른 ✕를 누르면 이전 pending은 교체된다.

_PENDING_CANCEL_TTL = 60.0


def _is_pending_cancel(target_id: str) -> bool:
    """target_id가 현재 취소 대기 중인지 반환. TTL 만료 시 자동 정리."""
    pending = st.session_state.get("_pending_cancel")
    if not pending:
        return False
    target, ts = pending
    if time.monotonic() - ts > _PENDING_CANCEL_TTL:
        st.session_state.pop("_pending_cancel", None)
        return False
    return target == target_id


def _mark_pending_cancel(target_id: str) -> None:
    """target_id를 취소 대기로 표시. 이전 pending은 자동 교체된다."""
    st.session_state._pending_cancel = (target_id, time.monotonic())


def _clear_pending_cancel() -> None:
    st.session_state.pop("_pending_cancel", None)


def persist_assignments():
    """배정을 파일에 저장한다. 실패 시 호출 전 스냅샷으로 롤백."""
    _backup = dict(st.session_state.assignments)
    try:
        save_assignments(assignments_file, st.session_state.assignments,
                         st.session_state.get("_assignments_mtime"))
        st.session_state._assignments_mtime = _file_mtime(assignments_file)
    except StaleFileError:
        st.session_state.assignments = _backup
        st.error("다른 사용자가 먼저 저장했습니다. 페이지를 새로고침하세요.")
        st.stop()
    except Exception:
        st.session_state.assignments = _backup
        raise


def persist_releases():
    """해제를 파일에 저장한다. 실패 시 호출 전 스냅샷으로 롤백."""
    _backup = dict(st.session_state.releases)
    try:
        save_releases(releases_file, st.session_state.releases,
                      st.session_state.get("_releases_mtime"))
        st.session_state._releases_mtime = _file_mtime(releases_file)
    except StaleFileError:
        st.session_state.releases = _backup
        st.error("다른 사용자가 먼저 저장했습니다. 페이지를 새로고침하세요.")
        st.stop()
    except Exception:
        st.session_state.releases = _backup
        raise
    st.session_state.released_slots = releases_to_slot_set(st.session_state.releases)


# _compute_auto_released는 assignment_status.compute_auto_released로 분리됨.
# session_state·module-level globals를 인자로 주입하는 thin wrapper만 남긴다.
def _compute_auto_released() -> dict[tuple[str, str, int], tuple[str, str]]:
    return compute_auto_released(
        st.session_state.assignments,
        requests,
        DATE_TO_DAY,
        DATE_TO_SHEET,
        DAY_TO_SHEETS,
    )


def log_audit(action: str, subject: str = "", details: dict | None = None):
    operator = (st.session_state.get("operator_name", "") or "").strip() or "unknown"
    append_audit_event(audit_file, operator, action, subject, details or {})


with st.sidebar:
    st.caption(f"배정 저장: {assignments_file.name}")
    if st.button("저장본 다시 불러오기"):
        _reload_session_data()
        loaded_count = len(st.session_state.assignments)
        log_audit("reload_assignments", details={"loaded_count": loaded_count})
        st.rerun()
    if st.button("엑셀 데이터 다시 불러오기"):
        cached_load.clear()
        st.rerun()


# 자동 해제 (ROOM_CHANGE → 원래 강의실) + 수동 해제 합산
auto_released = _compute_auto_released()  # {(sheet, room, period): subject_key}
all_released_slots = st.session_state.released_slots | set(auto_released.keys())

# ──────────────────────────────────────────────
# 탭
# ──────────────────────────────────────────────

def _split_assign_count(key):
    """ROOM_SPLIT 요청의 기존 배정 건수를 센다."""
    return sum(1 for k in st.session_state.assignments
               if k == key or k.startswith(key + "+"))

def _next_split_key(key):
    """ROOM_SPLIT 추가 배정용 키를 생성한다. 충돌 없는 다음 키 반환."""
    if key not in st.session_state.assignments:
        # +N 키만 남아있을 수도 있으므로 확인
        has_suffix = any(k.startswith(key + "+") for k in st.session_state.assignments)
        if not has_suffix:
            return key
    max_n = 1
    for k in st.session_state.assignments:
        if k.startswith(key + "+"):
            try:
                n = int(k[len(key) + 1:])
                max_n = max(max_n, n)
            except ValueError:
                pass
    return f"{key}+{max_n + 1}"

# ── 전체 진행 상황 (탭 간 공유) ──
_n_done = sum(1 for r in requests if compute_status(r, st.session_state.assignments, room_capacity) == "완료")
_n_todo = sum(1 for r in requests if compute_status(r, st.session_state.assignments, room_capacity) == "미배정")
_n_skip = sum(1 for r in requests if compute_status(r, st.session_state.assignments, room_capacity) == "미확정")
# 완료 = 자동(NORMAL_EXAM + NO_EXAM, 분류만으로 결정) + 수동(ROOM_CHANGE/SPLIT 배정 충족)
_n_auto_done = sum(
    1 for r in requests if r.category in (Category.NORMAL_EXAM, Category.NO_EXAM)
)
_n_manual_done = _n_done - _n_auto_done
# P열 직접 입력(배정 워크플로 미경유 수동 처리) — 프로그램 완료와 disjoint
_manual_entries = detect_manual_processed_entries(requests, st.session_state.assignments)
_n_manual_p = len(_manual_entries)
_n_manual_p_todo = sum(
    1 for req, _ in _manual_entries
    if req.category in (Category.ROOM_CHANGE, Category.ROOM_SPLIT)
)
_n_manual_p_skip = _n_manual_p - _n_manual_p_todo
_n_effective_done = _n_done + _n_manual_p
_review_rows = build_review_queue_rows(
    requests, timetable_data, room_capacity, st.session_state.assignments,
    all_released_slots,
)
_done_keys = {r.key for r in requests if compute_status(r, st.session_state.assignments, room_capacity) == "완료"}
_review_keys = {row["과목명"] for row in _review_rows}
_done_with_issues = len(_done_keys & _review_keys)
_done_help = f"이 중 {_done_with_issues}건이 검수 큐에 있습니다" if _done_with_issues else None

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "기존 시간표", "배정 작업", "배정 현황", "결과 검증", "통계",
])


# ── 탭 1: 기존 시간표 ──
with tab1:
    st.subheader("기존 시간표 (원본)")
    day1 = st.selectbox("일자", SHEET_ORDER, key="t1_day")

    raw = timetable_data.get(day1, {})
    room_data = {}
    color_map = {}
    for room, periods in raw.items():
        room_data[room] = {}
        for p, (val, rgb) in periods.items():
            room_data[room][p] = val
            color_map[(room, p)] = "#5F9EA0" if rgb == "FF5F9EA0" else "#D2691E"

    st.html(render_timetable_html(room_data, color_map, room_capacity))
    c1, c2 = st.columns(2)
    c1.markdown('<span style="background:#D2691E;color:#fff;padding:2px 8px;border-radius:3px;">갈색 = 수업</span>', unsafe_allow_html=True)
    c2.markdown('<span style="background:#5F9EA0;color:#fff;padding:2px 8px;border-radius:3px;">청록 = 특수/예약</span>', unsafe_allow_html=True)


# ── 탭 2: 배정 작업 ──
with tab2:
    st.subheader("배정 작업")

    # ── 진행 상황 요약 ──
    _pct = _n_effective_done / len(requests) if requests else 0.0
    if _n_manual_p:
        _ptext = (f"진행률 {int(_pct * 100)}% — 처리 {_n_effective_done} "
                  f"(프로그램 {_n_done} + P열수동 {_n_manual_p}) / 전체 {len(requests)}")
    else:
        _ptext = f"진행률 {int(_pct * 100)}% — 완료 {_n_done} / 전체 {len(requests)}"
    st.progress(_pct, text=_ptext)
    _m1, _m2, _m3, _m4, _m5 = st.columns(5)
    with _m1:
        st.metric("완료", f"{_n_done}건", help=_done_help)
        st.caption(f"자동 {_n_auto_done} / 수동 {_n_manual_done}")
    with _m2:
        st.metric("미배정", f"{_n_todo}건")
        if _n_manual_p_todo:
            st.caption(f"P열 직접 처리 {_n_manual_p_todo}")
    with _m3:
        st.metric(
            "미확정", f"{_n_skip}건",
            help="분류 단계에서 정보 부족으로 분류 안 된 항목. "
                 "탭3 검수 큐의 '세부' 컬럼에서 사유를 확인하고 "
                 "요청 엑셀을 정정한 뒤 사이드바의 '엑셀 데이터 다시 불러오기'를 누르면 재분류됩니다.",
        )
        if _n_manual_p_skip:
            st.caption(f"P열 직접 처리 {_n_manual_p_skip}")
    _m4.metric("검수 큐", f"{len(_review_rows)}건")
    _m5.metric("전체", f"{len(requests)}건")

    # ── 필터/검색 ──
    _fc1, _fc2, _fc3 = st.columns([1.5, 1, 2])
    with _fc1:
        u_cat_filter = st.multiselect(
            "분류 필터",
            options=list(CAT_LABELS.values()),
            default=list(CAT_LABELS.values()),
            key="u_cat_filter",
        )
    with _fc2:
        u_status_filter = st.multiselect(
            "상태 필터",
            options=["완료", "미배정", "미확정"],
            default=["미배정", "미확정"],
            key="u_status_filter",
        )
    with _fc3:
        u_search = st.text_input("검색 (과목명/교수명/학과)", key="u_search")

    inv_cat = {v: k for k, v in CAT_LABELS.items()}
    selected_cats = {inv_cat[l] for l in u_cat_filter}

    # ── 잔여 슬롯 히트맵 (접힌 상태) ──
    with st.expander("잔여 슬롯 히트맵"):
        u_heat_day = st.selectbox("일자", SHEET_ORDER, key="u_heat_day")
        raw_heat = timetable_data.get(u_heat_day, {})

        # 배정된 슬롯 수집 (히트맵 반영용)
        heat_assigned = set()
        for a in st.session_state.assignments.values():
            if a["sheet"] == u_heat_day:
                for ap in a["periods"]:
                    heat_assigned.add((a["room"], ap))

        thresholds = [("전체", 0), ("30석+", 30), ("50석+", 50), ("80석+", 80), ("100석+", 100)]
        heatmap_rows = []
        for label, min_c in thresholds:
            row = {"구간": label}
            for p in range(15):
                count = 0
                for room in raw_heat:
                    cap = room_capacity.get(room, 0)
                    if cap < min_c:
                        continue
                    is_released = (u_heat_day, room, p) in all_released_slots
                    if not is_released and p in raw_heat[room]:
                        continue
                    if (room, p) in heat_assigned:
                        continue
                    count += 1
                row[f"{p}교시"] = count
            heatmap_rows.append(row)

        df_heat = pd.DataFrame(heatmap_rows).set_index("구간")

        def color_cell(val):
            if isinstance(val, (int, float)):
                if val == 0:
                    return "background-color: #FF4444; color: white;"
                elif val <= 3:
                    return "background-color: #FF8888; color: white;"
                elif val <= 10:
                    return "background-color: #FFDD88;"
                else:
                    return "background-color: #88CC88;"
            return ""

        st.dataframe(df_heat.style.map(color_cell), **_STRETCH, height=220)
        st.caption("빨강=0~3개 (부족) / 노랑=4~10개 / 초록=11개+ (여유)")

        # ── 교시별 수급 현황 ──
        _demand = {p: 0 for p in range(15)}
        for req in requests:
            if req.category not in (Category.ROOM_CHANGE, Category.ROOM_SPLIT):
                continue
            if compute_status(req, st.session_state.assignments, room_capacity) == "완료":
                continue
            if not req.exam_date or req.exam_date not in DATE_TO_DAY:
                continue
            exam_day = DATE_TO_DAY[req.exam_date]
            if _resolve_sheet(req.exam_date, exam_day) != u_heat_day:
                continue
            for p in resolve_needed_periods(req, exam_day):
                _demand[p] += 1

        _supply_row = heatmap_rows[0]  # "전체" 구간
        _sd_rows = []
        _sd_supply = {"구분": "빈 강의실"}
        _sd_demand = {"구분": "미배정 수요"}
        _sd_diff = {"구분": "부족분"}
        for p in range(15):
            s = _supply_row[f"{p}교시"]
            d = _demand[p]
            _sd_supply[f"{p}교시"] = s
            _sd_demand[f"{p}교시"] = d
            _sd_diff[f"{p}교시"] = max(0, d - s)
        _sd_rows = [_sd_supply, _sd_demand, _sd_diff]
        df_sd = pd.DataFrame(_sd_rows).set_index("구분")

        def _color_sd(val):
            if isinstance(val, (int, float)) and val > 0:
                return "background-color: #FF4444; color: white;"
            return ""

        st.markdown("**교시별 수급 현황**")
        st.dataframe(df_sd.style.map(_color_sd, subset=pd.IndexSlice["부족분", :]),
                     **_STRETCH, height=140)

        # ── 교시별 빈 강의실 상세 ──
        with st.expander("빈 강의실 상세 보기"):
            _det_col1, _det_col2 = st.columns(2)
            with _det_col1:
                det_period = st.selectbox("교시", list(range(15)), key="heat_det_period")
            with _det_col2:
                det_min_cap = st.number_input("최소 수용인원", min_value=0, value=0, step=10,
                                              key="heat_det_cap")
            det_rows = []
            for room in sorted(raw_heat.keys()):
                cap = room_capacity.get(room, 0)
                if cap < det_min_cap:
                    continue
                is_released = (u_heat_day, room, det_period) in all_released_slots
                occupied = det_period in raw_heat.get(room, {})
                assigned = (room, det_period) in heat_assigned
                if assigned:
                    continue
                if occupied and not is_released:
                    continue
                status = "해제됨" if (occupied and is_released) else "빈 강의실"
                det_rows.append({"강의실": room, "수용인원": cap, "상태": status})
            if det_rows:
                st.dataframe(pd.DataFrame(det_rows), **_STRETCH, hide_index=True)
            else:
                st.warning("해당 조건의 빈 강의실이 없습니다.")

    # ── 미배정 섹션 ──
    def _is_split_assigned(key):
        """해당 키의 배정이 분반 모드인지 확인."""
        return any(st.session_state.assignments[k].get("category") == LABEL_ROOM_SPLIT
                   for k in st.session_state.assignments
                   if k == key or k.startswith(key + "+"))

    target_reqs = [r for r in requests
                   if r.category in (Category.ROOM_CHANGE, Category.ROOM_SPLIT)
                   and compute_status(r, st.session_state.assignments, room_capacity) == "미배정"]

    # 필터 적용된 미배정 목록
    filtered_targets = []
    for req in target_reqs:
        if req.category not in selected_cats:
            continue
        if "미배정" not in u_status_filter:
            continue
        if u_search:
            q = u_search.strip()
            if q not in req.name and q not in req.professor and q not in req.department and q not in req.key:
                continue
        filtered_targets.append(req)

    _active_sel_req = None  # 해제 UI 기본값용

    st.markdown(f"### 미배정 ({len(filtered_targets)}건)")
    if filtered_targets:
        target_rows = []
        for req in filtered_targets:
            _sc = _split_assign_count(req.key)
            target_rows.append({
                "분류": CAT_LABELS[req.category],
                "과목명": req.key,
                "수강생": req.students,
                "시험일": str(req.exam_date or ""),
                "시험교시": f"{req.exam_start}~{req.exam_end}" if req.exam_start is not None else "",
                "원래 강의실": req.room,
                "기배정": f"{_sc}건" if _sc else "",
                "요청사항": req.remarks,
            })
        st.dataframe(pd.DataFrame(target_rows), **_STRETCH, hide_index=True, height=200)

        # 배정 패널
        options = [f"{r.key} | {r.students}명 | {CAT_LABELS[r.category]}"
                   for r in filtered_targets]
        sel_idx = st.selectbox("과목 선택", range(len(options)),
                               format_func=lambda i: options[i], key="u_sel")
        sel_req = filtered_targets[sel_idx]
        _active_sel_req = sel_req

        left, right = st.columns([1.2, 0.8])
        with left:
            st.markdown(f"**{sel_req.key}**")
            st.write(f"수강생: {sel_req.students}명 | 원래 강의실: {sel_req.room}")
            if sel_req.exam_start is not None:
                st.write(f"시험일: {sel_req.exam_date} | 교시: {sel_req.exam_start}~{sel_req.exam_end}")
            if sel_req.schedule_raw:
                st.caption(f"수업시간표: {sel_req.schedule_raw.replace('~', '-')}")
            if sel_req.remarks:
                st.info(f"요청사항: {sel_req.remarks}")

            # 강의실 해제
            with st.expander("강의실 해제 (기존강의실 불필요)"):
                _rel_default_day_idx = 0
                _rel_default_room = None
                if sel_req.exam_date and sel_req.exam_date in DATE_TO_DAY:
                    _sel_day = DATE_TO_DAY[sel_req.exam_date]
                    _sel_sheet = _resolve_sheet(sel_req.exam_date, _sel_day)
                    if _sel_sheet in SHEET_ORDER:
                        _rel_default_day_idx = SHEET_ORDER.index(_sel_sheet)
                    _rel_default_room = sel_req.room

                _cur_rel_ctx = f"{_rel_default_day_idx}|{_rel_default_room}"
                if st.session_state.get("_rel_ctx") != _cur_rel_ctx:
                    st.session_state.pop("rel_day", None)
                    st.session_state.pop("rel_room", None)
                    st.session_state.pop("rel_periods", None)
                    st.session_state["_rel_ctx"] = _cur_rel_ctx

                rel_day = st.selectbox("일자", SHEET_ORDER, index=_rel_default_day_idx, key="rel_day")
                raw_rel = timetable_data.get(rel_day, {})

                # 시간표 외 강의실(수업시간표에만 있는)의 점유 슬롯 수집
                _rel_extra = collect_extra_room_slots(requests, rel_day, raw_rel)

                occupied_rooms = sorted(
                    set(r for r in raw_rel if raw_rel[r]) | set(_rel_extra.keys())
                )
                if occupied_rooms:
                    room_idx = 0
                    if _rel_default_room and _rel_default_room in occupied_rooms:
                        room_idx = occupied_rooms.index(_rel_default_room)
                    rel_room = st.selectbox("강의실", occupied_rooms, index=room_idx, key="rel_room")
                    # 이미 해제된 교시 제외
                    _rkey_check = f"{rel_day}|{rel_room}"
                    _already_released = set(
                        st.session_state.releases.get(_rkey_check, {}).get("periods", [])
                    )
                    # 교시 목록: 시간표 또는 수업시간표 슬롯
                    if rel_room in raw_rel:
                        occ_periods = sorted(p for p in raw_rel[rel_room].keys() if p not in _already_released)
                        period_labels = {p: f"{p}교시 ({raw_rel[rel_room][p][0][:20]})" for p in occ_periods}
                    else:
                        occ_periods = sorted(p for p in _rel_extra.get(rel_room, {}).keys() if p not in _already_released)
                        period_labels = {p: f"{p}교시 ({_rel_extra[rel_room][p][:20]})" for p in occ_periods}
                    if occ_periods:
                        rel_periods = st.multiselect("해제할 교시", occ_periods,
                                                     format_func=lambda p: period_labels[p], key="rel_periods")
                        if rel_periods and st.button("해제", key="rel_submit", type="primary"):
                            rkey = f"{rel_day}|{rel_room}"
                            existing = st.session_state.releases.get(rkey, {
                                "sheet": rel_day, "room": rel_room, "periods": []
                            })
                            merged = sorted(set(existing["periods"]) | set(rel_periods))
                            st.session_state.releases[rkey] = {
                                "sheet": rel_day, "room": rel_room, "periods": merged
                            }
                            persist_releases()
                            log_audit("release", f"{rel_day}/{rel_room}", {"periods": rel_periods})
                            st.rerun()
                    else:
                        st.info("해당 강의실에 수업이 없습니다.")
                else:
                    st.info("해당 일자에 수업이 있는 강의실이 없습니다.")

                if st.session_state.releases:
                    with st.expander(f"현재 해제 목록 ({len(st.session_state.releases)}건)"):
                        for rkey, rel in list(st.session_state.releases.items()):
                            _rc1, _rc2 = st.columns([4, 1])
                            period_str = ", ".join(str(p) for p in rel["periods"])
                            _rel_raw = timetable_data.get(rel["sheet"], {})
                            _rel_classes = []
                            for _rp in rel["periods"]:
                                _rv = _rel_raw.get(rel["room"], {}).get(_rp)
                                if _rv:
                                    _rel_classes.append(f"{_rp}교시:{_rv[0][:15]}")
                            _class_info = f" ({', '.join(_rel_classes)})" if _rel_classes else ""
                            _rel_cap = room_capacity.get(rel["room"], "")
                            _cap_str = f" [{_rel_cap}명]" if _rel_cap else ""
                            _rc1.write(f"{rel['sheet']} / {rel['room']}{_cap_str} / {period_str}교시{_class_info}")
                            _rel_target = f"release:{rkey}"
                            _rel_pending = _is_pending_cancel(_rel_target)
                            if _rc2.button(
                                "✓ 정말?" if _rel_pending else "취소",
                                key=f"rel_cancel_{rkey}",
                                type="primary" if _rel_pending else "secondary",
                                help="한 번 더 누르면 해제가 취소됩니다." if _rel_pending else None,
                            ):
                                if _rel_pending:
                                    del st.session_state.releases[rkey]
                                    persist_releases()
                                    log_audit("unrelease", rkey, {"room": rel["room"], "periods": rel["periods"]})
                                    _clear_pending_cancel()
                                else:
                                    _mark_pending_cancel(_rel_target)
                                st.rerun()

            # 미니 점유 현황 격자
            if sel_req.exam_date and sel_req.exam_date in DATE_TO_DAY:
                _grid_day = DATE_TO_DAY[sel_req.exam_date]
                _grid_sheet = _resolve_sheet(sel_req.exam_date, _grid_day)
                _grid_raw = timetable_data.get(_grid_sheet, {})
                _grid_periods = resolve_needed_periods(sel_req, _grid_day)
                if _grid_periods:
                    # 시간표 외 강의실(수업시간표에만 있는)의 점유 슬롯 수집
                    _grid_extra = collect_extra_room_slots(requests, _grid_sheet, _grid_raw)
                    _grid_html = render_availability_grid(
                        _grid_raw, room_capacity, _grid_periods, 0,
                        st.session_state.assignments, _grid_sheet, all_released_slots,
                        extra_slots=_grid_extra)
                    if _grid_html:
                        st.markdown(f"**강의실 점유 현황** ({_grid_sheet})")
                        st.html(_grid_html)
                        st.caption("◯ = 빈 교시 / 수업 = 점유 / 배정 = 기배정 / *강의실 = 시간표 외")

        with right:
            if sel_req.exam_date and sel_req.exam_date in DATE_TO_DAY:
                exam_day = DATE_TO_DAY[sel_req.exam_date]
                sheet = _resolve_sheet(sel_req.exam_date, exam_day)
                raw_sel = timetable_data.get(sheet, {})
                needed_periods = resolve_needed_periods(sel_req, exam_day)

                _existing_splits = _split_assign_count(sel_req.key)

                # 배정 모드 결정
                _keep_as_is = False
                _default_mode = 1 if sel_req.category == Category.ROOM_SPLIT else 0
                _mode = st.radio(
                    "배정 방식",
                    [
                        "이동: 다른 강의실로 옮김",
                        "분반: 원래 + 추가 강의실 동시 사용",
                        "유지: 원래 강의실 그대로",
                    ],
                    index=_default_mode,
                    key="u_assign_mode",
                    horizontal=True,
                )
                is_split = _mode.startswith("분반")
                if _mode.startswith("유지"):
                    _keep_as_is = True
                _save_cat = CAT_LABELS[sel_req.category]
                if is_split:
                    _save_cat = LABEL_ROOM_SPLIT

                if _keep_as_is:
                    orig_cap = room_capacity.get(sel_req.room, 0)
                    # 교시 폴백: needed_periods 없으면 슬롯 전체 교시 사용
                    _keep_periods = needed_periods
                    if not _keep_periods:
                        _keep_periods = sorted({
                            p for s in sel_req.slots if s.day == exam_day
                            for p in range(max(0, s.start), min(14, s.end) + 1)
                        })
                    if not _keep_periods:
                        st.warning("교시 정보가 없어 기존 강의실을 확정할 수 없습니다.")
                    else:
                        st.info(f"기존 강의실 {sel_req.room} ({orig_cap}명) 유지 — {_keep_periods}교시")
                    if st.button("기존 강의실로 확정", key="u_keep_assign", type="primary",
                                 disabled=not _keep_periods):
                        st.session_state.assignments[sel_req.key] = {
                            "room": sel_req.room,
                            "sheet": sheet,
                            "periods": _keep_periods,
                            "original_room": sel_req.room,
                            "students": sel_req.students,
                            "category": LABEL_NORMAL_EXAM,
                            "keep_orig": True,
                        }
                        persist_assignments()
                        log_audit("assign", sel_req.key, {
                            "sheet": sheet, "room": sel_req.room, "mode": "기존유지",
                        })
                        st.rerun()
                elif not needed_periods:
                    st.warning("시험 교시 정보가 없어 배정할 수 없습니다.")
                    free = []
                else:

                    if is_split:
                        _split_assignments = [
                            st.session_state.assignments[k]
                            for k in st.session_state.assignments
                            if k == sel_req.key or k.startswith(sel_req.key + "+")
                        ]
                        _existing_keep_orig = any(
                            bool(a.get("keep_orig", True))
                            for a in _split_assignments
                            if a.get("category") == LABEL_ROOM_SPLIT
                        )
                        if _existing_splits:
                            _keep_orig = _existing_keep_orig
                            st.checkbox(
                                "기존 강의실 유지",
                                value=_keep_orig,
                                key="u_keep_orig",
                                disabled=True,
                                help="첫 분반 배정의 유지/미유지 설정을 따릅니다.",
                            )
                        else:
                            _keep_orig = st.checkbox("기존 강의실 유지", value=True, key="u_keep_orig")
                        orig_cap = room_capacity.get(sel_req.room, 0)
                        if _existing_splits:
                            _split_rooms = [st.session_state.assignments[k]["room"]
                                            for k in st.session_state.assignments
                                            if k == sel_req.key or k.startswith(sel_req.key + "+")]
                            st.success(f"{'원래 강의실 ' + sel_req.room + ' (' + str(orig_cap) + '명) 유지' if _keep_orig else '기존 강의실 미유지 — 전체 이동'}\n"
                                       f"기배정 {_existing_splits}건: {', '.join(_split_rooms)}\n"
                                       f"— 추가 강의실 검색")
                        else:
                            if _keep_orig:
                                st.success(f"원래 강의실 {sel_req.room} ({orig_cap}명) 유지 — 추가 강의실 검색")
                            else:
                                st.info(f"기존 강의실 미유지 — 새 강의실로 분반 배정")
                        if not _keep_orig and _existing_splits:
                            _assigned_cap = 0
                            _seen_rooms = set()
                            for a in _split_assignments:
                                _room = str(a.get("room", ""))
                                if _room and _room not in _seen_rooms:
                                    _assigned_cap += _room_cap(room_capacity, _room)
                                    _seen_rooms.add(_room)
                            _remaining_cap = max(0, sel_req.students - _assigned_cap)
                            search_cap = max(1, _remaining_cap)
                            st.caption(f"현재 확보 좌석 {_assigned_cap} / 필요 {sel_req.students} (남은 {_remaining_cap})")
                        else:
                            search_cap = (sel_req.students + 1) // 2
                    else:
                        _keep_orig = False
                        search_cap = sel_req.students
                    st.markdown(f"**빈 강의실 ({sheet}, {needed_periods}교시, {search_cap}명+)**")
                    free = get_free_rooms(raw_sel, room_capacity, needed_periods,
                                          search_cap, st.session_state.assignments, sheet,
                                          all_released_slots)

                if free:
                    room_options = [f"{r} (수용 {c}명)" for r, c in free]
                    room_idx = st.selectbox("강의실 선택", range(len(room_options)),
                                            format_func=lambda i: room_options[i], key="u_room")
                    chosen_room = free[room_idx][0]
                    _chosen_cap = free[room_idx][1]
                    _cap_short = not is_split and _chosen_cap < search_cap
                    if _cap_short:
                        st.error(f"수용인원({_chosen_cap}명)이 수강생({sel_req.students}명)보다 적어 배정할 수 없습니다.")

                    if st.button("배정", key="u_assign", type="primary", disabled=_cap_short):
                        _ak = _next_split_key(sel_req.key) if is_split else sel_req.key
                        st.session_state.assignments[_ak] = {
                            "room": chosen_room,
                            "sheet": sheet,
                            "periods": needed_periods,
                            "original_room": sel_req.room,
                            "students": sel_req.students,
                            "category": _save_cat,
                            "keep_orig": _keep_orig if is_split else False,
                        }
                        persist_assignments()
                        log_audit("assign", _ak, {
                            "sheet": sheet, "room": chosen_room, "periods": needed_periods,
                            "mode": "분반" if is_split else "이동",
                        })
                        st.rerun()
                elif needed_periods:
                    _hint = (
                        f"{search_cap}명 이상 수용 가능한 빈 강의실이 없습니다.\n\n"
                        "다음을 시도해보세요:\n"
                        "- 아래 **조건 직접 검색**에서 최소 수용인원을 더 낮춤"
                    )
                    if is_split:
                        _hint += " (분반은 0도 가능 — 작은 강의실 여러 개로 분할)"
                    _hint += "\n- 교시 범위를 좁혀 검색 (예: 시험이 일부 교시만)"
                    st.info(_hint)

                with st.expander("조건 직접 검색"):
                    mc1, mc2 = st.columns(2)
                    m_period = mc1.slider("교시 범위", 0, 14,
                                          (needed_periods[0], needed_periods[-1]) if needed_periods else (4, 5),
                                          key="u_manual_p")
                    m_cap = mc2.number_input("최소 수용인원", min_value=0, value=0, step=10, key="u_manual_c")
                    m_periods = list(range(m_period[0], m_period[1] + 1))
                    m_free = get_free_rooms(raw_sel, room_capacity, m_periods, m_cap,
                                            st.session_state.assignments, sheet,
                                            all_released_slots)
                    if m_free:
                        st.dataframe(pd.DataFrame(m_free, columns=["강의실", "수용인원"]),
                                     **_STRETCH, hide_index=True)
                        m_room_options = [f"{r} (수용 {c}명)" for r, c in m_free]
                        m_room_idx = st.selectbox("강의실 선택", range(len(m_room_options)),
                                                  format_func=lambda i: m_room_options[i],
                                                  key="u_manual_room")
                        m_chosen = m_free[m_room_idx][0]
                        _m_chosen_cap = m_free[m_room_idx][1]
                        _m_cap_short = not is_split and _m_chosen_cap < sel_req.students
                        if _m_cap_short:
                            st.error(f"수용인원({_m_chosen_cap}명)이 수강생({sel_req.students}명)보다 적어 배정할 수 없습니다.")
                        if needed_periods and m_periods != needed_periods:
                            st.warning(f"선택 교시({m_periods[0]}~{m_periods[-1]})가 "
                                       f"요구 교시({needed_periods[0]}~{needed_periods[-1]})와 다릅니다.")
                        if st.button("배정", key="u_manual_assign", type="primary", disabled=_m_cap_short):
                            _ak = _next_split_key(sel_req.key) if is_split else sel_req.key
                            st.session_state.assignments[_ak] = {
                                "room": m_chosen,
                                "sheet": sheet,
                                "periods": m_periods,
                                "original_room": sel_req.room,
                                "students": sel_req.students,
                                "category": _save_cat,
                                "keep_orig": _keep_orig if is_split else False,
                            }
                            persist_assignments()
                            log_audit("assign", _ak, {
                                "sheet": sheet, "room": m_chosen, "periods": m_periods,
                                "mode": "분반" if is_split else "이동",
                            })
                            st.rerun()
                    else:
                        _cap_part = f", {m_cap}명+" if m_cap else ""
                        st.warning(
                            f"선택 조건({m_period[0]}~{m_period[-1]}교시{_cap_part})에 "
                            "맞는 빈 강의실이 없습니다.\n\n"
                            "다음 중 하나를 시도해보세요:\n"
                            "- 최소 수용인원을 더 낮춰서 다시 검색 "
                            "(분반은 0으로도 가능 — 작은 강의실 여러 개로 분할)\n"
                            "- 교시 범위를 더 좁혀서 다시 검색 "
                            "(예: 2~3교시 → 2교시만)\n"
                            "- 그래도 안 되면 시험 시간 자체를 요청자와 협의"
                        )
            else:
                st.warning("시험일자 정보가 없어 자동 검색이 불가합니다.")
    elif target_reqs:
        st.info("필터 조건에 맞는 미배정 항목 없음")
    else:
        st.success("모든 과목 배정 완료!")

    # ── 배정 작업 현황 (분류별 + 개별 취소) ──
    with st.expander("배정 작업 현황", expanded=bool(st.session_state.assignments)):
        if st.session_state.assignments:
            # 분류별 그룹핑
            _by_cat: dict[str, list[tuple[str, dict]]] = {}
            for k, a in st.session_state.assignments.items():
                cat = a.get("category", "기타")
                _by_cat.setdefault(cat, []).append((k, a))

            for cat, items in _by_cat.items():
                st.markdown(f"**{cat}** ({len(items)}건)")
                for key, a in items:
                    _lc1, _lc2, _lc3 = st.columns([3, 2, 0.5])
                    _lc1.write(f"**{key}** ({a['students']}명)")
                    _lc2.write(f"{a['original_room']} → {a['room']} | {a['sheet']} {a['periods']}교시")
                    _cancel_target = f"assignment:{key}"
                    _cancel_pending = _is_pending_cancel(_cancel_target)
                    if _lc3.button(
                        "✓ 정말?" if _cancel_pending else "✕",
                        key=f"log_cancel_{key}",
                        type="primary" if _cancel_pending else "secondary",
                        help="한 번 더 누르면 배정이 삭제됩니다." if _cancel_pending else None,
                    ):
                        if _cancel_pending:
                            del st.session_state.assignments[key]
                            persist_assignments()
                            log_audit("unassign", key, {"room": a["room"]})
                            _clear_pending_cancel()
                        else:
                            _mark_pending_cancel(_cancel_target)
                        st.rerun()
                st.markdown("---")

            st.caption(f"총 {len(st.session_state.assignments)}건 배정")
        else:
            st.info("배정 내역이 없습니다.")

    # ── P열 직접 입력 내역 (배정 워크플로 미경유 수동 처리, 읽기 전용) ──
    with st.expander(f"P열 직접 입력 내역 ({_n_manual_p}건)"):
        st.caption(
            "배정 작업을 거치지 않고 요청 엑셀 P열(요청사항 처리)에 직접 강의실을 적은 건 — "
            "원본 엑셀 값이라 읽기 전용입니다. 프로그램 배정과 합쳐 실질 결과물이 됩니다."
        )
        if _manual_entries:
            _mp_rows = [{
                "과목명": req.name,
                "분반": req.ban,
                "학생수": req.students,
                "분류": CAT_LABELS[req.category],
                "P열 강의실": ", ".join(rooms),
                "시험일": str(req.exam_date) if req.exam_date else "",
            } for req, rooms in _manual_entries]
            st.dataframe(pd.DataFrame(_mp_rows), **_STRETCH, hide_index=True, height=260)
        else:
            st.info("P열에 직접 입력된 건이 없습니다.")


# ── 탭 3: 배정 현황 ──
with tab3:
    st.subheader("배정 현황")

    # ── 진행 상황 요약 ──
    _pct = _n_effective_done / len(requests) if requests else 0.0
    if _n_manual_p:
        _ptext = (f"진행률 {int(_pct * 100)}% — 처리 {_n_effective_done} "
                  f"(프로그램 {_n_done} + P열수동 {_n_manual_p}) / 전체 {len(requests)}")
    else:
        _ptext = f"진행률 {int(_pct * 100)}% — 완료 {_n_done} / 전체 {len(requests)}"
    st.progress(_pct, text=_ptext)
    _s1, _s2, _s3, _s4, _s5 = st.columns(5)
    with _s1:
        st.metric("완료", f"{_n_done}건", help=_done_help)
        st.caption(f"자동 {_n_auto_done} / 수동 {_n_manual_done}")
    with _s2:
        st.metric("미배정", f"{_n_todo}건")
        if _n_manual_p_todo:
            st.caption(f"P열 직접 처리 {_n_manual_p_todo}")
    with _s3:
        st.metric(
            "미확정", f"{_n_skip}건",
            help="분류 단계에서 정보 부족으로 분류 안 된 항목. "
                 "아래 검수 큐의 '세부' 컬럼에서 사유를 확인하고 "
                 "요청 엑셀을 정정한 뒤 사이드바의 '엑셀 데이터 다시 불러오기'를 누르면 재분류됩니다.",
        )
        if _n_manual_p_skip:
            st.caption(f"P열 직접 처리 {_n_manual_p_skip}")
    _s4.metric("검수 큐", f"{len(_review_rows)}건")
    _s5.metric("전체", f"{len(requests)}건")

    # ── 필터/검색 ──
    _sc1, _sc2, _sc3 = st.columns([1.5, 1, 2])
    with _sc1:
        s_cat_filter = st.multiselect(
            "분류 필터",
            options=list(CAT_LABELS.values()),
            default=list(CAT_LABELS.values()),
            key="s_cat_filter",
        )
    with _sc2:
        s_status_filter = st.multiselect(
            "상태 필터",
            options=["완료", "미배정", "미확정"],
            default=["완료", "미배정", "미확정"],
            key="s_status_filter",
        )
    with _sc3:
        s_search = st.text_input("검색 (과목명/교수명/학과)", key="s_search")

    # ── 검수 큐 ──
    _filtered_review = [r for r in _review_rows
                        if r["분류"] in s_cat_filter
                        and (not s_search or s_search.strip() in r["과목명"])]
    st.markdown(f"### 검수 큐 ({len(_filtered_review)}건)")
    if _filtered_review:
        df_review = pd.DataFrame(_filtered_review)
        st.dataframe(df_review, **_STRETCH, hide_index=True, height=220)
        buf_review = io.BytesIO()
        df_review.to_excel(buf_review, index=False, engine="openpyxl")
        st.download_button("검수 큐 엑셀 다운로드", buf_review.getvalue(), "검수큐.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.success("검토 필요 항목 없음")

    # ── 완료 섹션 ──
    st.markdown("---")
    s_inv_cat = {v: k for k, v in CAT_LABELS.items()}
    s_selected_cats = {s_inv_cat[l] for l in s_cat_filter}

    completed_rows = []
    for req in requests:
        status = compute_status(req, st.session_state.assignments, room_capacity)
        if status != "완료":
            continue
        if req.category not in s_selected_cats:
            continue
        if "완료" not in s_status_filter:
            continue
        if s_search:
            q = s_search.strip()
            if q not in req.name and q not in req.professor and q not in req.department and q not in req.key:
                continue

        _assign_keys = [k for k in st.session_state.assignments
                        if k == req.key or k.startswith(req.key + "+")]
        if _assign_keys:
            _rooms = [st.session_state.assignments[k]["room"] for k in _assign_keys]
            _first = st.session_state.assignments[_assign_keys[0]]
            completed_rows.append({
                "상태": f"배정 완료 ({len(_assign_keys)}건)" if len(_assign_keys) > 1 else "배정 완료",
                "분류": CAT_LABELS[req.category],
                "과목명": req.key,
                "수강생": req.students,
                "시험일": _first["sheet"],
                "강의실": f"{_first['original_room']} → {', '.join(_rooms)}",
                "요청사항": req.remarks,
            })
        elif req.category == Category.NORMAL_EXAM:
            completed_rows.append({
                "상태": "기존 확정",
                "분류": CAT_LABELS[req.category],
                "과목명": req.key,
                "수강생": req.students,
                "시험일": str(req.exam_date or ""),
                "강의실": req.room,
                "요청사항": req.remarks,
            })
        elif req.category == Category.NO_EXAM:
            completed_rows.append({
                "상태": "시험 미실시",
                "분류": CAT_LABELS[req.category],
                "과목명": req.key,
                "수강생": req.students,
                "시험일": "",
                "강의실": "",
                "요청사항": req.remarks,
            })

    st.markdown(f"### 완료 ({len(completed_rows)}건)")
    if completed_rows:
        df_completed = pd.DataFrame(completed_rows)
        st.dataframe(df_completed, **_STRETCH, hide_index=True, height=300)
        _buf_done = io.BytesIO()
        df_completed.to_excel(_buf_done, index=False, engine="openpyxl")
        st.download_button("완료 내역 엑셀 다운로드", _buf_done.getvalue(), "완료내역.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("완료 항목 없음 (필터 확인)")

    # ── 내보내기 + 작업 이력 ──
    with st.expander("내보내기 + 작업 이력"):
        # ── 요청 엑셀에 배정 결과(P열) 채워 다운로드 ──
        # 배정 결과를 '요청사항 처리'(P열)에 기입한 원본 양식 사본. 이후 사용자가
        # 미확정 행을 직접 입력하고 → 결과 검증 탭에서 충돌을 최종 점검한다.
        st.markdown("**요청사항 처리(P열) 채운 요청 엑셀**")
        st.caption("원본 요청 엑셀 양식 그대로, P열에 프로그램 배정 결과를 기입한 사본입니다. "
                   "배정·시험·미실시로 확정된 행만 채우고 미확정 행은 비워 둬 직접 입력할 수 있습니다. "
                   "원본 파일은 바뀌지 않습니다.")
        if st.button("P열 채운 요청 엑셀 만들기", key="u_gen_p_export"):
            st.session_state._p_export_bytes = generate_request_with_processed(
                str(req_file), requests, st.session_state.assignments)
        if st.session_state.get("_p_export_bytes"):
            st.download_button(
                "다운로드 (요청_P열채움.xlsx)",
                data=st.session_state._p_export_bytes,
                file_name=f"{req_file.stem}_P열채움.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        st.markdown("---")

        if st.session_state.assignments:
            export_rows = []
            for key, a in st.session_state.assignments.items():
                export_rows.append({
                    "과목": key,
                    "분류": a["category"],
                    "수강생": a["students"],
                    "원래 강의실": a["original_room"],
                    "배정 강의실": a["room"],
                    "일자": a["sheet"],
                    "교시": str(a["periods"]),
                })
            df_assigned = pd.DataFrame(export_rows)
            st.dataframe(df_assigned, **_STRETCH, hide_index=True)

            c1, c2, c3 = st.columns(3)
            csv = df_assigned.to_csv(index=False).encode("utf-8-sig")
            c1.download_button("CSV 다운로드", csv, "배정결과.csv", "text/csv")

            buf = io.BytesIO()
            df_assigned.to_excel(buf, index=False, engine="openpyxl")
            c2.download_button("엑셀 다운로드", buf.getvalue(), "배정결과.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            _confirm_clear = c3.checkbox("초기화 확인", key="u_confirm_clear",
                                          help="체크 후 초기화 버튼을 누르세요")
            if c3.button("전체 초기화", type="secondary", disabled=not _confirm_clear):
                cleared_count = len(st.session_state.assignments)
                cleared_releases = len(st.session_state.releases)
                st.session_state.assignments = {}
                st.session_state.releases = {}
                st.session_state.released_slots = set()
                persist_assignments()
                persist_releases()
                log_audit("clear_all", details={"count": cleared_count, "releases": cleared_releases})
                st.rerun()
        else:
            st.info("배정 내역이 없습니다.")

        st.markdown("---")
        st.markdown("**작업 이력 (최근 100건)**")
        audit_events = read_audit_events(audit_file, limit=100)
        if audit_events:
            audit_rows = []
            for e in audit_events:
                audit_rows.append({
                    "시각(UTC)": e.get("timestamp", ""),
                    "작업자": e.get("operator", ""),
                    "동작": e.get("action", ""),
                    "과목": e.get("subject", ""),
                    "세부": format_audit_details(e.get("details")),
                })
            st.dataframe(pd.DataFrame(audit_rows), **_STRETCH, hide_index=True, height=220)
        else:
            st.caption("이력이 없습니다.")


# ── 탭 4: 요청 시간표 + 충돌 감지 ──
with tab4:
    st.subheader("결과 검증 (충돌 감지)")

    # ── 원본 '요청사항 처리'(P열) 강의실 배정 점검 (읽기 전용 감사) ──
    # 요청 엑셀 P열에 사람이 미리 적은 배정 결정을 감사한다. 두 종류를 본다:
    #   (1) P열↔P열  — 같은 시험일자·강의실에 교시가 겹치는 다른 과목 (이중 배정)
    #   (2) P열↔시간표 — P열 강의실이 기존 수업시간표 점유와 겹침 (해제·자기수업 제외)
    # 프로그램 내부 배정(_assignments.json)·수업시간표 파생 강의실과는 별개.
    _proc_conflicts, _proc_unjudged = detect_processed_room_conflicts(requests)
    _tt_overlaps = detect_timetable_overlaps(
        requests, timetable_data, DATE_TO_SHEET, DATE_TO_DAY, all_released_slots,
        occupant_index=data.get("course_exam_index"))
    with st.container(border=True):
        st.markdown("**원본 '요청사항 처리'(P열) 강의실 점검**")
        st.caption("요청 엑셀 P열에 미리 적힌 배정 결정 감사 — 같은 과목 분반끼리는 충돌이 아닙니다. "
                   "해제(자동/수동)되었거나 자기 강의실에서 보는 시험은 제외합니다.")

        # (1) P열 ↔ P열 이중 배정
        if _proc_conflicts:
            st.error(f"① 강의실 이중 배정 {len(_proc_conflicts)}건 — 같은 날·강의실·교시에 다른 과목")
            _df_pc = pd.DataFrame([{
                "시험일자": c.exam_date.isoformat(),
                "강의실": c.room,
                "과목 A": f"{c.req_a.name}({c.req_a.ban}) "
                          f"{c.req_a.exam_start}~{c.req_a.exam_end}교시",
                "행 A": c.req_a.row,
                "과목 B": f"{c.req_b.name}({c.req_b.ban}) "
                          f"{c.req_b.exam_start}~{c.req_b.exam_end}교시",
                "행 B": c.req_b.row,
            } for c in _proc_conflicts])
            st.dataframe(_df_pc, hide_index=True, **_STRETCH)
        else:
            st.success("① 강의실 이중 배정 없음")

        # (2) P열 ↔ 기존 수업시간표 겹침
        if _tt_overlaps:
            _ov_groups = {}
            for o in _tt_overlaps:
                g = _ov_groups.setdefault(
                    (o.exam_date, o.room, o.req.row),
                    {"req": o.req, "room": o.room, "date": o.exam_date,
                     "periods": [], "occupants": []})
                g["periods"].append(o.period)
                if o.occupant not in g["occupants"]:
                    g["occupants"].append(o.occupant)
            st.warning(f"② 기존 수업과 겹침 {len(_ov_groups)}건 — P열 강의실이 그 시간 시간표에 점유됨 (확인 필요)")
            _df_ov = pd.DataFrame([{
                "시험일자": g["date"].isoformat(),
                "강의실": g["room"],
                "과목": f"{g['req'].name}({g['req'].ban})",
                "겹친 교시": ", ".join(f"{p}교시" for p in sorted(g["periods"])),
                "기존 점유(시간표)": ", ".join(g["occupants"]),
                "행": g["req"].row,
            } for g in _ov_groups.values()])
            st.dataframe(_df_ov, hide_index=True, **_STRETCH)
        else:
            st.success("② 기존 수업시간표와 겹침 없음")

        # 정보 부족으로 판정 불가한 행 (강의실은 있으나 날짜/교시 누락)
        if _proc_unjudged:
            with st.expander(
                f"확인 필요 — 강의실은 배정됐으나 날짜/교시 누락 {len(_proc_unjudged)}건"
            ):
                st.caption("아래 행은 정보 부족으로 충돌 판정에서 제외됨 — 원본 K·L·M열 보완 필요")
                _df_un = pd.DataFrame([{
                    "강의실": u.room,
                    "과목": f"{u.req.name}({u.req.ban})",
                    "사유": u.reason,
                    "행": u.req.row,
                } for u in _proc_unjudged])
                st.dataframe(_df_un, hide_index=True, **_STRETCH)
    st.divider()

    day3 = st.selectbox("일자", SHEET_ORDER, key="t3_day")

    full_data, full_color, full_tip, conflict_cells = _build_day_verification(
        day3, requests, timetable_data, st.session_state.assignments,
        auto_released, st.session_state.released_slots)

    conflict_count = len(conflict_cells)
    if conflict_count:
        st.error(f"충돌 {conflict_count}건 감지! 빨간 셀을 확인하세요.")
        with st.expander("충돌 상세 보기"):
            for room, p in sorted(conflict_cells):
                tip = full_tip.get((room, p), "")
                st.markdown(f"**{room} / {p}교시**")
                for line in tip.split("\n"):
                    st.write(line)
                st.markdown("---")
    else:
        st.success("충돌 없음")

    st.html(render_timetable_html(full_data, full_color, room_capacity, tooltip_map=full_tip))

    cols = st.columns(8)
    legend = [
        ("#D2691E", "기존 수업"), ("#5F9EA0", "특수/예약"),
        ("#4472C4", "시험/이동/분반"), ("#FF69B4", "미실시/해제"), ("#FFD700", "변경 요청(미배정)"),
        ("#9370DB", "분반 요청(미배정)"), ("#228B22", "시간표 외 강의실"), ("#FF0000", "충돌!"),
    ]
    for col, (c, label) in zip(cols, legend):
        fg = "#fff"
        col.markdown(f'<span style="background:{c};color:{fg};padding:2px 8px;'
                     f'border-radius:3px;">{label}</span>', unsafe_allow_html=True)

    # ── 엑셀 내보내기 (전체 일자, 시트별) ──
    def _generate_verification_excel():
        wb = _openpyxl.Workbook()
        wb.remove(wb.active)

        hdr_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        hdr_align = Alignment(horizontal="center")

        for day in SHEET_ORDER:
            data, colors, _, _ = _build_day_verification(
                day, requests, timetable_data, st.session_state.assignments,
                auto_released, st.session_state.released_slots)

            ws = wb.create_sheet(title=day)
            headers = ["강의실", "수용"] + [f"{p}교시 ({8 + p}~{9 + p}시)" for p in range(15)]
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align

            for ri, room in enumerate(sorted(data.keys()), 2):
                ws.cell(row=ri, column=1, value=room).font = Font(bold=True)
                ws.cell(row=ri, column=2, value=room_capacity.get(room, "")).alignment = hdr_align
                periods = data.get(room, {})
                for p in range(15):
                    cell = ws.cell(row=ri, column=p + 3, value=periods.get(p, ""))
                    hex_c = colors.get((room, p), "#FFFFFF")
                    rgb = hex_c.lstrip("#")
                    cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
                    if hex_c in DARK_BG:
                        cell.font = Font(color="FFFFFF")

            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 6
            for ci in range(3, 18):
                ws.column_dimensions[get_column_letter(ci)].width = 16
            ws.freeze_panes = "C2"

        # 범례 시트
        ws_legend = wb.create_sheet(title="범례")
        ws_legend.column_dimensions["A"].width = 6
        ws_legend.column_dimensions["B"].width = 25
        ws_legend.column_dimensions["C"].width = 40
        ws_legend.cell(row=1, column=1, value="색상").font = Font(bold=True)
        ws_legend.cell(row=1, column=2, value="구분").font = Font(bold=True)
        ws_legend.cell(row=1, column=3, value="설명").font = Font(bold=True)
        legend_items = [
            ("D2691E", "기존 수업", "시간표에 등록된 일반 수업"),
            ("5F9EA0", "특수/예약", "시간표에 등록된 특수/예약 슬롯"),
            ("4472C4", "시험/이동/분반", "시험 진행, 강의실 이동 배정, 분반 배정"),
            ("FF69B4", "미실시/해제", "시험 미실시 또는 수동 해제된 슬롯"),
            ("FFD700", "변경 요청(미배정)", "강의실 변경 요청 중 아직 미배정"),
            ("9370DB", "분반 요청(미배정)", "강의실 분반 요청 중 아직 미배정"),
            ("228B22", "시간표 외 강의실", "시간표에 없지만 수업시간표에 등록된 강의실의 수업 점유"),
            ("FF0000", "충돌", "같은 강의실+교시에 2건 이상 겹침"),
        ]
        for ri, (rgb, label, desc) in enumerate(legend_items, 2):
            cell = ws_legend.cell(row=ri, column=1, value="")
            cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
            ws_legend.cell(row=ri, column=2, value=label)
            ws_legend.cell(row=ri, column=3, value=desc)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    st.download_button(
        "결과 검증 엑셀 다운로드 (전체 일자)",
        data=_generate_verification_excel(),
        file_name="결과검증.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── 탭 5: 통계 ──
with tab5:
    st.subheader("통계")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**분류별 건수**")
        cat_counts = Counter(CAT_LABELS[r.category] for r in requests)
        df_cat = pd.DataFrame([{"분류": k, "건수": v} for k, v in cat_counts.items()])
        st.bar_chart(df_cat.set_index("분류"))

    with c2:
        st.markdown("**일별 시험 수**")
        day_counts = Counter()
        for r in requests:
            if r.category == Category.NORMAL_EXAM and r.exam_date in DATE_TO_DAY:
                sheet = _resolve_sheet(r.exam_date)
                day_counts[sheet] += 1
        rows = [{"일자": s, "시험 수": day_counts.get(s, 0)} for s in SHEET_ORDER]
        st.bar_chart(pd.DataFrame(rows).set_index("일자"))

    st.markdown("**강의실 가동률 (교시 점유율, 해제 반영)**")
    util_rows = []
    for sn in SHEET_ORDER:
        raw = timetable_data.get(sn, {})
        total = len(raw) * 15
        # 가동률 계산은 timetable에 있는 강의실만 분자·분모로 다룬다.
        # 시간표 외 강의실(수업시간표에는 있지만 timetable 엑셀에 없는
        # 강의실)의 release는 분자 차감에서 제외 — 그렇지 않으면
        # filled가 음수가 될 수 있다.
        released_count = sum(
            1 for s, r, p in all_released_slots if s == sn and r in raw
        )
        filled = sum(len(periods) for periods in raw.values()) - released_count
        pct = round(filled / total * 100, 1) if total else 0
        util_rows.append({"일자": sn, "점유율(%)": pct, "사용": filled, "해제": released_count, "전체": total})
    st.dataframe(pd.DataFrame(util_rows), **_STRETCH, hide_index=True)

    st.markdown("**시험주간 통산 미사용 강의실 (전 기간 수업·배정 0건)**")
    _unused = compute_unused_rooms(
        timetable_data, room_capacity, st.session_state.assignments, SHEET_ORDER
    )
    if _unused:
        df_unused = pd.DataFrame(
            [{"강의실": r, "수용인원": (c or "-")} for r, c in _unused]
        )
        st.dataframe(df_unused, **_STRETCH, hide_index=True, height=300)
        st.caption(
            f"총 {len(_unused)}개 — 시험 기간 내내 기존 수업·배정이 0건인 강의실. "
            "여유 공급(추가 배정 가능)으로 활용할 수 있다. 수용인원 '-'는 엑셀에 인원 미기재."
        )
    else:
        st.info("시험주간 내내 비어 있는 강의실이 없습니다.")

    st.markdown("---")
    m1, m2, m3, m4, m5 = st.columns(5)
    total = len(requests)
    m1.metric("총 요청", f"{total}건")
    m2.metric(LABEL_NORMAL_EXAM, f"{cat_counts.get(LABEL_NORMAL_EXAM, 0)}건")
    m3.metric("미실시", f"{cat_counts.get('미실시/대체과제', 0)}건")
    m4.metric("변경 요청", f"{cat_counts.get('강의실 변경', 0)}건")
    m5.metric("분반 요청", f"{cat_counts.get('강의실 분반', 0)}건")
