"""강의실 배정 데이터 로더

엑셀 파일 2개(요청사항 + 타임테이블)를 읽어
구조화된 데이터로 변환한다. 시각화/분류 전용 — 자동 배정 로직 없음.
"""

import re
import datetime
from dataclasses import dataclass, field
from enum import Enum, auto
from typing import Optional

import openpyxl

# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────

DAY_TO_SHEET = {
    "월": "4.27.(월)",
    "화": "4.21.(화)",
    "수": "4.22.(수)",
    "목": "4.23.(목)",
    "금": "4.24.(금)",
}

DATE_TO_DAY = {
    datetime.date(2026, 4, 21): "화",
    datetime.date(2026, 4, 22): "수",
    datetime.date(2026, 4, 23): "목",
    datetime.date(2026, 4, 24): "금",
    datetime.date(2026, 4, 27): "월",
}

PERIOD_TO_COL = {p: p + 6 for p in range(15)}

SCHEDULE_RE = re.compile(r"([월화수목금토일])\s*(\d+)(?:\s*~\s*(\d+))?\s*\(\s*([^)]*?)\s*\)")
_SHEET_NAME_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.\(([월화수목금토일])\)")

# 공백을 모두 제거한 정규화 형태로만 등록. 매칭 시 remarks도 동일하게
# 정규화되므로 "미실시"·"미 실시"·"미  실시" 등 띄어쓰기 변형을 모두
# 흡수한다.
NO_EXAM_KEYWORDS = [
    "미실시", "미시행", "대체과제", "과제대체",
    "미사용", "사용안함", "이용안함",
    "불필요", "필요없음", "온라인시험",
]

# room_choice 정규화 상수 — 비교용. 사용자가 적은 원본은 req.room_choice에
# 그대로 보존되어 화면 표시·로그에 쓰인다.
_RC_ROOM_CHANGE = "강의실변경요청"
_RC_ROOM_SPLIT = "강의실분반요청"
_RC_AS_IS = "기존강의실"


def _normalize_choice(s) -> str:
    """공백(다중공백, 탭, 줄바꿈 포함)을 모두 제거한 비교용 문자열을 반환."""
    if not s:
        return ""
    return re.sub(r"\s+", "", str(s))

SHEET_ORDER = ["4.21.(화)", "4.22.(수)", "4.23.(목)", "4.24.(금)", "4.27.(월)"]


# ──────────────────────────────────────────────
# 데이터 클래스
# ──────────────────────────────────────────────

@dataclass
class ScheduleSlot:
    day: str
    start: int
    end: int
    room: str


class Category(Enum):
    NORMAL_EXAM = auto()
    NO_EXAM = auto()
    ROOM_CHANGE = auto()
    ROOM_SPLIT = auto()
    SKIP = auto()


@dataclass
class ExamRequest:
    row: int
    department: str
    name: str
    ban: str
    professor: str
    students: int
    schedule_raw: str
    slots: list
    room: str
    exam_date: Optional[datetime.date]
    exam_start: Optional[int]
    exam_end: Optional[int]
    room_choice: Optional[str]
    remarks: str
    processed_room: str = ""
    key: str = ""
    category: Category = Category.SKIP
    skip_reason: str = ""

    def __post_init__(self):
        self.key = f"{self.name}-{self.ban}"


# ──────────────────────────────────────────────
# 파싱
# ──────────────────────────────────────────────

def parse_schedule(raw: str) -> list:
    if not raw or not raw.strip():
        return []
    slots = []
    for m in SCHEDULE_RE.finditer(raw):
        day = m.group(1)
        start = int(m.group(2))
        end = int(m.group(3)) if m.group(3) else start
        room = m.group(4).strip()
        slots.append(ScheduleSlot(day, start, end, room))
    return slots


def _parse_date(val) -> Optional[datetime.date]:
    if val is None or val == "":
        return None
    if isinstance(val, datetime.date) and not isinstance(val, datetime.datetime):
        if val.year < 2000:
            return None
        return val
    if isinstance(val, datetime.datetime):
        if val.year < 2000:
            return None
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        if val == "0000-01-01" or val == "":
            return None
        try:
            return datetime.datetime.strptime(val, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _parse_int(val) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def build_mappings_from_sheets(sheet_names: list, year: int) -> tuple:
    """시간표 시트 이름에서 날짜/요일/시트 매핑을 동적 생성한다.

    Returns:
        (date_to_day, day_to_sheet, sheet_order, date_to_sheet, day_to_sheets)
        * day_to_sheet   : 요일→시트(마지막 시트). 하위 호환용 단일 매핑.
        * day_to_sheets  : 요일→시트 목록(시간순). 다주차 안전 1:N 매핑.
        * date_to_sheet  : 날짜→시트. 1:1, 항상 안전.
    """
    date_to_day = {}
    day_to_sheet = {}
    day_to_sheets: dict[str, list[str]] = {}
    parsed = []

    for name in sheet_names:
        m = _SHEET_NAME_RE.match(name)
        if not m:
            continue
        month, day, dayname = int(m.group(1)), int(m.group(2)), m.group(3)
        try:
            dt = datetime.date(year, month, day)
        except ValueError:
            continue
        parsed.append((dt, dayname, name))

    parsed.sort(key=lambda x: x[0])
    date_to_sheet = {}
    sheet_order = []
    for dt, dayname, name in parsed:
        date_to_day[dt] = dayname
        day_to_sheet[dayname] = name  # 동일 요일 시 마지막 시트 (하위 호환)
        day_to_sheets.setdefault(dayname, []).append(name)  # 같은 요일 모든 시트
        date_to_sheet[dt] = name      # 날짜→시트 직접 매핑 (다주차 안전)
        sheet_order.append(name)

    return date_to_day, day_to_sheet, sheet_order, date_to_sheet, day_to_sheets


def _infer_year(requests: list):
    """요청 목록에서 첫 번째 유효한 시험일자의 연도를 추출한다."""
    for req in requests:
        if req.exam_date is not None:
            return req.exam_date.year
    return None


def _has_no_exam_keyword(remarks: str) -> bool:
    if not remarks:
        return False
    normalized = _normalize_choice(remarks)
    return any(kw in normalized for kw in NO_EXAM_KEYWORDS)


# ──────────────────────────────────────────────
# 요청 로딩 및 분류
# ──────────────────────────────────────────────

def load_requests(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    requests = []

    for row_num in range(2, ws.max_row + 1):
        r = lambda c: ws.cell(row=row_num, column=c).value

        if r(4) != "공통":
            continue

        schedule_raw = str(r(9) or "")
        slots = parse_schedule(schedule_raw)
        remarks = str(r(15) or "")

        req = ExamRequest(
            row=row_num,
            department=str(r(3) or ""),
            name=str(r(5) or "").strip(),
            ban=str(r(6) or "").strip(),
            professor=str(r(7) or ""),
            students=_parse_int(r(8)) or 0,
            schedule_raw=schedule_raw,
            slots=slots,
            room=str(r(10) or "").strip(),
            exam_date=_parse_date(r(11)),
            exam_start=_parse_int(r(12)),
            exam_end=_parse_int(r(13)),
            room_choice=str(r(14)).strip() if r(14) else None,
            remarks=remarks,
            processed_room=str(r(16) or "").strip(),
        )
        requests.append(req)

    # 중복 키 감지 → 행 번호 붙여 고유화
    seen: dict[str, int] = {}
    for req in requests:
        if req.key in seen:
            seen[req.key] += 1
            req.key = f"{req.key}#{req.row}"
        else:
            seen[req.key] = 1
    # 첫 번째 등장도 충돌이었으면 고유화
    duped = {k for k, cnt in seen.items() if cnt > 1}
    for req in requests:
        if req.key in duped:
            req.key = f"{req.key}#{req.row}"

    wb.close()
    return requests


def normalize_ban(ban) -> str:
    """분반 비교용 정규화 — 숫자 분반의 앞 0을 제거('01'→'1'). 영숫자('A0')는 원형 유지.

    시간표 점유자 텍스트('과목명-01')와 요청 시트 분반('1' 또는 '01')의 표기
    차이를 흡수한다.
    """
    b = str(ban or "").strip()
    if b.isdigit():
        return b.lstrip("0") or "0"
    return b


def load_course_exam_index(filepath: str) -> dict:
    """요청 엑셀의 모든 시트(공통+전공)에서 (교과목명, 분반)별 시험 메타를 모은다.

    ② 시간표 겹침 검사에서 "시간표 점유자가 그 시험주간에 실제로 그 방에서
    시험을 보는가"를 판정하는 데 쓴다. 정상 수업 시간표는 시험주간 점유의
    근거가 아니므로(시험만 점유), 점유자의 시험 일자/교시/미실시 여부로 검증한다.

    Returns:
        ``{(교과목명, 정규화분반): {exam_date, exam_start, exam_end, no_exam}}``
        — 공통이 두 시트에 중복되면 첫 등장만 유지.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    index: dict = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if str(ws.cell(row=1, column=5).value or "") != "교과목명":
            continue  # 요청 시트가 아닌 시트 건너뜀
        for row_num in range(2, ws.max_row + 1):
            r = lambda c: ws.cell(row=row_num, column=c).value
            name = str(r(5) or "").strip()
            if not name:
                continue
            key = (name, normalize_ban(r(6)))
            if key in index:
                continue
            index[key] = {
                "exam_date": _parse_date(r(11)),
                "exam_start": _parse_int(r(12)),
                "exam_end": _parse_int(r(13)),
                "no_exam": _has_no_exam_keyword(str(r(15) or "")),
            }
    wb.close()
    return index


def classify_requests(requests: list, date_to_day: dict = None) -> list:
    if date_to_day is None:
        date_to_day = DATE_TO_DAY
    for req in requests:
        if req.room_choice:
            req.room_choice = req.room_choice.strip()
        # 비교는 공백 변형을 흡수한 정규화 문자열로 한다. 원본은 표시·로그용으로 보존.
        _rc_norm = _normalize_choice(req.room_choice)

        if not req.slots:
            req.category = Category.SKIP
            req.skip_reason = "수업시간표(I열) 누락 — I열에 시간표 입력 필요"
            continue

        if all(s.room == "" for s in req.slots):
            req.category = Category.SKIP
            req.skip_reason = "수업시간표 강의실 표기 누락 — I열의 (강의실) 부분 정정 필요"
            continue

        # 시험일 요일에 해당하는 강의실로 재설정
        if req.exam_date is not None and req.exam_date in date_to_day:
            exam_day = date_to_day[req.exam_date]
            for slot in req.slots:
                if slot.day == exam_day and slot.room:
                    req.room = slot.room
                    break

        if _rc_norm in (_RC_ROOM_CHANGE, _RC_ROOM_SPLIT):
            if req.exam_date is not None and req.exam_date in date_to_day:
                req.category = (Category.ROOM_CHANGE
                                if _rc_norm == _RC_ROOM_CHANGE
                                else Category.ROOM_SPLIT)
            else:
                req.category = Category.SKIP
                _action = '변경' if _rc_norm == _RC_ROOM_CHANGE else '분반'
                req.skip_reason = (
                    f"강의실 {_action} 요청이지만 시험일자(K열) 없음/범위 밖 "
                    f"({req.exam_date}) — K열 정정 필요"
                )
            continue

        if req.exam_date is None:
            if _has_no_exam_keyword(req.remarks):
                req.category = Category.NO_EXAM
            else:
                req.category = Category.SKIP
                req.skip_reason = (
                    "시험일자(K열) 없고 미실시 키워드도 없음 — "
                    "K열 추가 또는 O열에 '미실시'/'대체과제' 등 키워드 추가"
                )
            continue

        if req.exam_date not in date_to_day:
            req.category = Category.SKIP
            req.skip_reason = (
                f"시험일자 범위 밖 ({req.exam_date}) — "
                "K열 정정 또는 시험 기간 확인 필요"
            )
            continue

        if req.exam_start is not None and req.exam_end is not None:
            req.category = Category.NORMAL_EXAM
        elif _rc_norm == _RC_AS_IS:
            req.category = Category.NORMAL_EXAM
        else:
            if _has_no_exam_keyword(req.remarks):
                req.category = Category.NO_EXAM
            else:
                req.category = Category.NORMAL_EXAM

    return requests


# ──────────────────────────────────────────────
# 타임테이블 로딩
# ──────────────────────────────────────────────

def load_timetable(filepath: str) -> tuple:
    """Returns: (room_to_row, room_capacity, timetable_data)

    timetable_data: {sheet_name: {room_code: {period: (cell_value, color_rgb)}}}
    """
    wb = openpyxl.load_workbook(filepath)

    room_to_row = {}
    room_capacity = {}
    timetable_data = {}

    for name in wb.sheetnames:
        ws = wb[name]
        room_to_row[name] = {}
        timetable_data[name] = {}
        for row_num in range(2, ws.max_row + 1):
            room_code = ws.cell(row=row_num, column=2).value
            if room_code:
                room_code = str(room_code).strip()
                room_to_row[name][room_code] = row_num
                cap = ws.cell(row=row_num, column=4).value
                if cap and room_code not in room_capacity:
                    try:
                        room_capacity[room_code] = int(cap)
                    except (ValueError, TypeError):
                        pass
                periods = {}
                for p in range(15):
                    cell = ws.cell(row=row_num, column=PERIOD_TO_COL[p])
                    val = cell.value
                    if val is not None and str(val).strip():
                        rgb = None
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                            rgb = cell.fill.fgColor.rgb
                        periods[p] = (str(val).strip(), rgb)
                timetable_data[name][room_code] = periods

    wb.close()
    return room_to_row, room_capacity, timetable_data


# ──────────────────────────────────────────────
# 통합 진입점
# ──────────────────────────────────────────────

def load_all(request_file: str, timetable_file: str) -> dict:
    """엑셀 2개를 읽어 전체 데이터를 반환."""
    requests = load_requests(request_file)
    course_exam_index = load_course_exam_index(request_file)
    room_to_row, room_capacity, timetable_data = load_timetable(timetable_file)

    # 시트 이름에서 날짜 매핑을 동적 생성 (파싱 실패 시 하드코딩 폴백)
    date_to_sheet = {}
    day_to_sheets: dict[str, list[str]] = {}
    year = _infer_year(requests)
    if year is not None:
        date_to_day, day_to_sheet, sheet_order, date_to_sheet, day_to_sheets = \
            build_mappings_from_sheets(list(timetable_data.keys()), year)
    if year is None or not sheet_order:
        date_to_day, day_to_sheet, sheet_order = DATE_TO_DAY, DAY_TO_SHEET, SHEET_ORDER
        # 하드코딩 폴백에서도 date_to_sheet/day_to_sheets 생성
        date_to_sheet = {}
        day_to_sheets = {d: [s] for d, s in day_to_sheet.items()}

    requests = classify_requests(requests, date_to_day)

    return {
        "requests": requests,
        "course_exam_index": course_exam_index,
        "room_to_row": room_to_row,
        "room_capacity": room_capacity,
        "timetable_data": timetable_data,
        "date_to_day": date_to_day,
        "day_to_sheet": day_to_sheet,
        "day_to_sheets": day_to_sheets,
        "date_to_sheet": date_to_sheet,
        "sheet_order": sheet_order,
    }
