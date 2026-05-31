"""data_loader 단위 테스트."""

import datetime
import pytest

from data_loader import (
    _parse_date,
    parse_schedule,
    classify_requests,
    build_mappings_from_sheets,
    _infer_year,
    ScheduleSlot,
    Category,
)


class TestParseDate:
    def test_datetime(self):
        assert _parse_date(datetime.datetime(2026, 4, 21)) == datetime.date(2026, 4, 21)

    def test_date(self):
        assert _parse_date(datetime.date(2026, 4, 21)) == datetime.date(2026, 4, 21)

    def test_date_old_year(self):
        assert _parse_date(datetime.date(1, 1, 1)) is None

    def test_string_valid(self):
        assert _parse_date("2026-04-22") == datetime.date(2026, 4, 22)

    def test_string_zero(self):
        assert _parse_date("0000-01-01") is None

    def test_empty(self):
        assert _parse_date("") is None
        assert _parse_date(None) is None

    def test_datetime_old_year(self):
        assert _parse_date(datetime.datetime(1, 1, 1)) is None


class TestParseSchedule:
    def test_single_period(self):
        result = parse_schedule("수1(K107)")
        assert len(result) == 1
        assert result[0] == ScheduleSlot("수", 1, 1, "K107")

    def test_range(self):
        result = parse_schedule("화4~5(K106)")
        assert result[0].start == 4
        assert result[0].end == 5

    def test_multi_slot(self):
        result = parse_schedule("화4~5(K106), 목4(K106)")
        assert len(result) == 2
        assert result[1].day == "목"
        assert result[1].end == 4

    def test_empty_room(self):
        result = parse_schedule("월2~3()")
        assert result[0].room == ""

    def test_empty_string(self):
        assert parse_schedule("") == []
        assert parse_schedule(None) == []


def _make_req(**kwargs):
    from data_loader import ExamRequest
    defaults = dict(
        row=1, department="테스트", name="테스트과목", ban="01",
        professor="교수", students=30, schedule_raw="화2~3(K106)",
        slots=[ScheduleSlot("화", 2, 3, "K106")], room="K106",
        exam_date=datetime.date(2026, 4, 21), exam_start=2, exam_end=3,
        room_choice="기존 강의실", remarks="",
    )
    defaults.update(kwargs)
    return ExamRequest(**defaults)


class TestClassify:
    def test_normal_exam(self):
        req = _make_req()
        classify_requests([req])
        assert req.category == Category.NORMAL_EXAM

    def test_no_exam_keyword(self):
        req = _make_req(exam_date=None, remarks="중간고사 미실시")
        classify_requests([req])
        assert req.category == Category.NO_EXAM

    def test_skip_no_schedule(self):
        req = _make_req(slots=[], schedule_raw="")
        classify_requests([req])
        assert req.category == Category.SKIP

    def test_room_change_with_date(self):
        req = _make_req(room_choice="강의실 변경 요청")
        classify_requests([req])
        assert req.category == Category.ROOM_CHANGE

    def test_room_change_no_date_is_skip(self):
        req = _make_req(room_choice="강의실 변경 요청", exam_date=None)
        classify_requests([req])
        assert req.category == Category.SKIP

    def test_room_split_with_date(self):
        req = _make_req(room_choice="강의실 분반 요청")
        classify_requests([req])
        assert req.category == Category.ROOM_SPLIT

    def test_trailing_space(self):
        req = _make_req(room_choice="강의실 변경 요청 ")
        classify_requests([req])
        assert req.category == Category.ROOM_CHANGE

    def test_leading_space(self):
        req = _make_req(room_choice=" 강의실 분반 요청")
        classify_requests([req])
        assert req.category == Category.ROOM_SPLIT

    def test_room_choice_no_internal_space(self):
        """내부 공백 없는 변형도 동일하게 인식."""
        req = _make_req(room_choice="강의실변경요청")
        classify_requests([req])
        assert req.category == Category.ROOM_CHANGE

    def test_room_choice_partial_space(self):
        """일부 공백만 있는 변형도 흡수."""
        req = _make_req(room_choice="강의실 분반요청")
        classify_requests([req])
        assert req.category == Category.ROOM_SPLIT

    def test_room_choice_multi_space(self):
        """다중 공백/탭/줄바꿈도 흡수."""
        req = _make_req(room_choice="강의실  변경\t요청")
        classify_requests([req])
        assert req.category == Category.ROOM_CHANGE

    def test_room_choice_as_is_no_space(self):
        """'기존 강의실'도 공백 변형 흡수."""
        # exam_start/end가 있으면 그것만으로 NORMAL_EXAM이 되므로
        # _rc_norm 경로를 타도록 둘 다 None으로 세팅
        req = _make_req(room_choice="기존강의실",
                        exam_start=None, exam_end=None)
        classify_requests([req])
        assert req.category == Category.NORMAL_EXAM

    def test_no_exam_keyword_with_space_variants(self):
        """NO_EXAM 키워드의 공백 변형도 모두 흡수."""
        for variant in ("미실시", "미 실시", "미  실시", "대체 과제",
                        "사용 안함", "온라인 시험"):
            req = _make_req(exam_date=None, remarks=f"비고: {variant}")
            classify_requests([req])
            assert req.category == Category.NO_EXAM, f"실패한 변형: {variant!r}"

    def test_room_choice_original_preserved(self):
        """비교는 정규화로 하지만 원본 room_choice 텍스트는 보존된다."""
        original = "강의실변경요청"
        req = _make_req(room_choice=original)
        classify_requests([req])
        assert req.room_choice == original
        assert req.category == Category.ROOM_CHANGE


class TestKeyDedup:
    """중복 키 감지 회귀 테스트."""

    def test_unique_keys_unchanged(self):
        from data_loader import ExamRequest
        reqs = [_make_req(row=2, name="과목A", ban="01"), _make_req(row=3, name="과목B", ban="01")]
        # load_requests의 중복 감지 로직 재현
        seen = {}
        for req in reqs:
            if req.key in seen:
                seen[req.key] += 1
                req.key = f"{req.key}#{req.row}"
            else:
                seen[req.key] = 1
        duped = {k for k, cnt in seen.items() if cnt > 1}
        for req in reqs:
            if req.key in duped:
                req.key = f"{req.key}#{req.row}"

        assert reqs[0].key == "과목A-01"
        assert reqs[1].key == "과목B-01"

    def test_duplicate_keys_get_row_suffix(self):
        from data_loader import ExamRequest
        reqs = [_make_req(row=2, name="과목A", ban="01"), _make_req(row=5, name="과목A", ban="01")]
        seen = {}
        for req in reqs:
            if req.key in seen:
                seen[req.key] += 1
                req.key = f"{req.key}#{req.row}"
            else:
                seen[req.key] = 1
        duped = {k for k, cnt in seen.items() if cnt > 1}
        for req in reqs:
            if req.key in duped:
                req.key = f"{req.key}#{req.row}"

        assert reqs[0].key == "과목A-01#2"
        assert reqs[1].key == "과목A-01#5"
        assert reqs[0].key != reqs[1].key


class TestBuildMappings:
    """동적 날짜 매핑 생성 테스트."""

    def test_basic_mapping(self):
        sheets = ["4.21.(화)", "4.22.(수)", "4.23.(목)", "4.24.(금)", "4.27.(월)"]
        d2d, d2s, order, *_ = build_mappings_from_sheets(sheets, 2026)
        assert d2d[datetime.date(2026, 4, 21)] == "화"
        assert d2d[datetime.date(2026, 4, 27)] == "월"
        assert d2s["화"] == "4.21.(화)"
        assert order == sheets  # 날짜순 정렬

    def test_different_year(self):
        sheets = ["10.15.(수)", "10.16.(목)"]
        d2d, d2s, order, *_ = build_mappings_from_sheets(sheets, 2025)
        assert d2d[datetime.date(2025, 10, 15)] == "수"
        assert d2s["목"] == "10.16.(목)"
        assert len(order) == 2

    def test_non_matching_sheets_ignored(self):
        sheets = ["4.21.(화)", "summary", "기타"]
        d2d, d2s, order, *_ = build_mappings_from_sheets(sheets, 2026)
        assert len(d2d) == 1
        assert len(order) == 1

    def test_classify_with_dynamic_mapping(self):
        """다른 학기 날짜도 동적 매핑으로 올바르게 분류."""
        sheets = ["10.15.(수)"]
        d2d, *_ = build_mappings_from_sheets(sheets, 2025)
        req = _make_req(exam_date=datetime.date(2025, 10, 15))
        classify_requests([req], date_to_day=d2d)
        assert req.category == Category.NORMAL_EXAM

    def test_classify_out_of_range_with_dynamic_mapping(self):
        """동적 매핑 범위 밖 날짜는 SKIP."""
        sheets = ["10.15.(수)"]
        d2d, *_ = build_mappings_from_sheets(sheets, 2025)
        req = _make_req(exam_date=datetime.date(2025, 12, 1))
        classify_requests([req], date_to_day=d2d)
        assert req.category == Category.SKIP

    def test_day_to_sheets_multi_week(self):
        """동일 요일이 여러 주에 걸쳐 있으면 day_to_sheets는 모두 보존하고,
        day_to_sheet(하위 호환)는 마지막 시트만 가진다."""
        sheets = ["4.15.(화)", "4.16.(수)", "4.22.(화)", "4.23.(수)"]
        d2d, d2s_single, order, d2sheet, d2sheets = \
            build_mappings_from_sheets(sheets, 2026)

        # 1:N 매핑은 같은 요일의 모든 시트를 시간순으로 보존
        assert d2sheets["화"] == ["4.15.(화)", "4.22.(화)"]
        assert d2sheets["수"] == ["4.16.(수)", "4.23.(수)"]

        # 1:1 매핑(하위 호환)은 마지막 시트만 보존 (기존 동작 유지)
        assert d2s_single["화"] == "4.22.(화)"
        assert d2s_single["수"] == "4.23.(수)"

        # 날짜→시트는 항상 안전
        assert d2sheet[datetime.date(2026, 4, 15)] == "4.15.(화)"
        assert d2sheet[datetime.date(2026, 4, 22)] == "4.22.(화)"

    def test_day_to_sheets_single_week(self):
        """단일 주차에서는 day_to_sheets와 day_to_sheet가 동등."""
        sheets = ["4.21.(화)", "4.22.(수)", "4.23.(목)"]
        _, d2s_single, _, _, d2sheets = build_mappings_from_sheets(sheets, 2026)
        for day, sheet in d2s_single.items():
            assert d2sheets[day] == [sheet]


class TestInferYear:
    def test_from_requests(self):
        req = _make_req(exam_date=datetime.date(2025, 10, 15))
        assert _infer_year([req]) == 2025

    def test_none_when_no_dates(self):
        req = _make_req(exam_date=None, remarks="미실시")
        assert _infer_year([req]) is None


class TestLoadProcessedRoom:
    def test_reads_col16(self, tmp_path):
        """load_requests가 16번째 열(요청사항 처리)을 processed_room으로 읽는다."""
        import openpyxl
        from data_loader import load_requests
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["변경요청", "순번", "소속", "과목구분", "교과목명", "분반",
                   "담당교수", "수강생수", "수업시간표", "강의실", "시험일자",
                   "시작교시", "종료교시", "강의실선택", "요청사항", "요청사항 처리"]
        ws.append(headers)  # 로더는 2행부터 읽음
        ws.append([0, 1, "학부", "공통", "사이버윤리", "03", "교수", 78,
                   "수2~3(M213)", "M213", "2026-06-17", 6, 7, None,
                   "과제대체", "N212"])
        path = tmp_path / "req.xlsx"
        wb.save(path)

        reqs = load_requests(str(path))
        assert len(reqs) == 1
        assert reqs[0].processed_room == "N212"
        assert reqs[0].name == "사이버윤리"

    def test_processed_room_defaults_empty(self):
        req = _make_req()
        assert req.processed_room == ""


class TestNormalizeBan:
    def test_strip_leading_zero(self):
        from data_loader import normalize_ban
        assert normalize_ban("01") == "1"
        assert normalize_ban("1") == "1"
        assert normalize_ban("10") == "10"
        assert normalize_ban("0") == "0"
        assert normalize_ban(1) == "1"

    def test_alnum_and_empty(self):
        from data_loader import normalize_ban
        assert normalize_ban("A0") == "A0"
        assert normalize_ban(None) == ""
        assert normalize_ban("  02 ") == "2"


class TestLoadCourseExamIndex:
    def test_reads_all_sheets(self, tmp_path):
        import openpyxl
        from data_loader import load_course_exam_index
        hdr = ["", "", "", "과목구분", "교과목명", "분반", "", "", "", "",
               "시험일자", "시작교시", "종료교시", "", "요청사항", ""]
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "전공개설과목신청리스트 (공통)"
        ws1.append(hdr)
        ws1.append([None, None, None, "공통", "법학개론", "01", None, None, None,
                    None, "2026-06-16", 4, 5, None, "비고", None])
        ws2 = wb.create_sheet("전공개설과목신청리스트")
        ws2.append(hdr)
        ws2.append([None, None, None, "전공", "한국어어휘론", "01", None, None, None,
                    None, "2026-06-09", 2, 2, None, "시험주간 강의실 사용 안함", None])
        path = tmp_path / "req.xlsx"
        wb.save(path)

        idx = load_course_exam_index(str(path))
        assert idx[("법학개론", "1")]["exam_date"] == datetime.date(2026, 6, 16)
        assert idx[("법학개론", "1")]["exam_start"] == 4
        assert idx[("법학개론", "1")]["no_exam"] is False
        assert idx[("한국어어휘론", "1")]["exam_date"] == datetime.date(2026, 6, 9)
        assert idx[("한국어어휘론", "1")]["no_exam"] is True


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
